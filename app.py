# -*- coding: utf-8 -*-
"""
문서 비서 📄  (Naver Table OCR → 토지 등기 "표제부" + "갑구" 정리)
===============================================================

✅ 사용자 요구 반영(v0.6.0)
1) 표제부가 여러 지번(여러 페이지)에 따로 존재하면, **지번별로 표를 분리**해서 보여주고/엑셀 시트도 분리
2) 갑구도 **지번별로 분리**해서 정리
   - 컬럼: 순위번호, 등기목적, 접수, 등기원인, 권리자 및 기타사항
   - 권리자 및 기타사항이 여러 셀/여러 줄로 나뉘는 경우 → 같은 순위번호로 이어붙이기

기술 스택
- 네이버 CLOVA OCR General + enableTableDetection=true (표추출 OCR)
- 표( tables/cells )는 네이버가 준 rowIndex/columnIndex/Span을 기반으로 복원
- 표제부/갑구는 "작은 온톨로지(aliases + 정규화 규칙)"로 헤더 인식 및 컬럼 매핑

주의
- 표 추출은 네이버 콘솔 Domain 설정에서 "표 추출 여부"가 ON이어야 동작합니다.
- OCR이 표를 잘못 쪼개면 완벽 복원은 불가능하지만, 등기 표제부/갑구는 구조가 비교적 규칙적이라 성공률이 높습니다.

requirements.txt (권장)
----------------------
streamlit
requests
pandas
openpyxl
pypdf
PyPDF2
"""

from __future__ import annotations

import io
import json
import time
import uuid
import hashlib
import re
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# 0) 앱 설정
# ============================================================
APP_TITLE = "DOVI 📄 dev — 토지 등기(표제부+갑구+을구) 정리"
APP_VERSION = "v0.7.0"

DEFAULT_PASSWORD = "alohomora"  # 데모용
MAX_PAGES_PER_REQUEST = 10      # Naver General OCR PDF 최대 10페이지/요청
MAX_SHEETNAME_LEN = 31


# ============================================================
# 1) PDF 유틸
# ============================================================
def _import_pypdf():
    try:
        from pypdf import PdfReader, PdfWriter  # type: ignore
        return PdfReader, PdfWriter
    except Exception:
        from PyPDF2 import PdfReader, PdfWriter  # type: ignore
        return PdfReader, PdfWriter


def get_pdf_total_pages(pdf_bytes: bytes) -> int:
    PdfReader, _ = _import_pypdf()
    reader = PdfReader(io.BytesIO(pdf_bytes))
    return len(reader.pages)


def split_pdf_into_chunks(pdf_bytes: bytes, chunk_size: int) -> List[Tuple[bytes, int, int]]:
    """
    PDF bytes를 chunk_size 페이지 단위로 쪼개서
    [(chunk_pdf_bytes, start_page(1-indexed), end_page(1-indexed)), ...] 반환
    """
    PdfReader, PdfWriter = _import_pypdf()
    reader = PdfReader(io.BytesIO(pdf_bytes))
    total = len(reader.pages)

    out: List[Tuple[bytes, int, int]] = []
    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        w = PdfWriter()
        for i in range(start, end):
            w.add_page(reader.pages[i])
        buf = io.BytesIO()
        w.write(buf)
        out.append((buf.getvalue(), start + 1, end))
    return out


# ============================================================
# 2) 네이버 OCR 호출 (표추출 ON)
# ============================================================
def call_naver_ocr_table(
    file_bytes: bytes,
    api_url: str,
    secret_key: str,
    *,
    lang: str = "ko",
    timeout: int = 180,
) -> Dict[str, Any]:
    request_json = {
        "images": [{"format": "pdf", "name": "upload"}],
        "requestId": str(uuid.uuid4()),
        "version": "V2",
        "timestamp": int(round(time.time() * 1000)),
        "lang": lang,
        "enableTableDetection": True,
    }

    payload = {"message": json.dumps(request_json, ensure_ascii=False)}
    headers = {"X-OCR-SECRET": secret_key}
    files = {"file": ("upload.pdf", file_bytes, "application/pdf")}

    try:
        r = requests.post(api_url, headers=headers, data=payload, files=files, timeout=timeout)
        ok = r.status_code == 200
        j = None
        if "application/json" in (r.headers.get("Content-Type") or ""):
            try:
                j = r.json()
            except Exception:
                j = None
        return {"ok": ok, "status_code": r.status_code, "text": (r.text or "")[:2000], "json": j}
    except Exception as e:
        return {"ok": False, "status_code": None, "error": str(e), "text": ""}


# ============================================================
# 3) OCR 테이블 파싱 + 페이지 텍스트라인(갑/을 구 판별용)
# ============================================================
@dataclass
class ParsedCell:
    row: int
    col: int
    row_span: int
    col_span: int
    text: str
    bbox: Optional[Tuple[float, float, float, float]] = None
    conf: Optional[float] = None


@dataclass
class ParsedTable:
    table_id: str
    sheet_name: str
    page_no: int
    table_index_on_page: int
    bbox: Optional[Tuple[float, float, float, float]]
    n_rows: int
    n_cols: int
    grid: List[List[str]]  # [r][c] (top-left only if merged)
    merges: List[Tuple[int, int, int, int]]  # (r0,c0,r1,c1) 0-based inclusive
    cells: List[ParsedCell]


@dataclass
class PageLine:
    y: float
    text: str


def _bbox_from_vertices(vertices: List[Dict[str, Any]]) -> Optional[Tuple[float, float, float, float]]:
    if not vertices:
        return None
    try:
        xs = [float(v.get("x", 0)) for v in vertices]
        ys = [float(v.get("y", 0)) for v in vertices]
        return (min(xs), min(ys), max(xs), max(ys))
    except Exception:
        return None


def _cell_text(cell: Dict[str, Any]) -> str:
    lines = cell.get("cellTextLines") or []
    out_lines: List[str] = []
    for ln in lines:
        words = ln.get("cellWords") or []
        wtxt = " ".join(
            (w.get("inferText") or "").strip()
            for w in words
            if (w.get("inferText") or "").strip()
        )
        wtxt = re.sub(r"\s+", " ", wtxt).strip()
        if wtxt:
            out_lines.append(wtxt)
    return "\n".join(out_lines).strip()


def _fields_to_page_lines(img: Dict[str, Any], *, y_thresh: float = 14.0) -> List[PageLine]:
    """
    OCR fields(일반 텍스트 박스)를 y기준으로 줄 단위로 묶어서 반환.
    (갑구/을구 구분을 위해 테이블 근처의 텍스트를 찾는 용도)
    """
    fields = img.get("fields") or []
    items = []
    for f in fields:
        txt = (f.get("inferText") or "").strip()
        if not txt:
            continue
        verts = ((f.get("boundingPoly") or {}).get("vertices")) or []
        if not verts:
            continue
        x0 = float(verts[0].get("x", 0))
        y0 = float(verts[0].get("y", 0))
        items.append((y0, x0, txt))

    if not items:
        return []

    items.sort(key=lambda x: (x[0], x[1]))
    lines: List[PageLine] = []
    cur = []
    last_y = items[0][0]

    def flush():
        nonlocal cur
        if not cur:
            return
        cur.sort(key=lambda x: x[1])
        y_avg = sum(x[0] for x in cur) / len(cur)
        text = " ".join(x[2] for x in cur).strip()
        text = re.sub(r"\s+", " ", text)
        if text:
            lines.append(PageLine(y=y_avg, text=text))
        cur = []

    for y0, x0, txt in items:
        if cur and abs(y0 - last_y) > y_thresh:
            flush()
        cur.append((y0, x0, txt))
        last_y = y0

    flush()
    return lines


def parse_tables_and_lines_from_ocr_json(
    ocr_json: Dict[str, Any], *, page_numbers: List[int]
) -> Tuple[List[ParsedTable], Dict[int, List[PageLine]]]:
    """
    반환:
      - tables: ParsedTable 리스트
      - page_lines: {page_no: [PageLine, ...]} (갑/을 구 판별용)
    """
    tables_out: List[ParsedTable] = []
    page_lines: Dict[int, List[PageLine]] = {}

    images = ocr_json.get("images", []) if isinstance(ocr_json, dict) else []

    for img_idx, img in enumerate(images):
        page_no = page_numbers[img_idx] if img_idx < len(page_numbers) else (img_idx + 1)

        # page lines
        page_lines[page_no] = _fields_to_page_lines(img)

        # tables
        tables = img.get("tables") or []
        if not tables:
            continue

        for t_idx, t in enumerate(tables, start=1):
            cells_raw = t.get("cells") or []
            parsed_cells: List[ParsedCell] = []

            max_r = 0
            max_c = 0
            merges: List[Tuple[int, int, int, int]] = []

            t_bbox = _bbox_from_vertices(((t.get("boundingPoly") or {}).get("vertices")) or [])

            for c in cells_raw:
                r = int(c.get("rowIndex", 0))
                col = int(c.get("columnIndex", 0))
                rspan = int(c.get("rowSpan", 1) or 1)
                cspan = int(c.get("columnSpan", 1) or 1)

                txt = _cell_text(c)
                c_bbox = _bbox_from_vertices(((c.get("boundingPoly") or {}).get("vertices")) or [])

                conf = c.get("inferConfidence")
                try:
                    conf_val = float(conf) if conf is not None else None
                except Exception:
                    conf_val = None

                parsed_cells.append(
                    ParsedCell(row=r, col=col, row_span=rspan, col_span=cspan, text=txt, bbox=c_bbox, conf=conf_val)
                )

                max_r = max(max_r, r + rspan)
                max_c = max(max_c, col + cspan)

                if rspan > 1 or cspan > 1:
                    merges.append((r, col, r + rspan - 1, col + cspan - 1))


            # 테이블 bbox가 응답에 없을 수 있어 cell bbox로 보강
            if t_bbox is None:
                xs0, ys0, xs1, ys1 = [], [], [], []
                for pc in parsed_cells:
                    if pc.bbox:
                        x0, y0, x1, y1 = pc.bbox
                        xs0.append(x0); ys0.append(y0); xs1.append(x1); ys1.append(y1)
                if xs0 and ys0 and xs1 and ys1:
                    t_bbox = (min(xs0), min(ys0), max(xs1), max(ys1))

            n_rows = max_r
            n_cols = max_c

            grid = [["" for _ in range(n_cols)] for _ in range(n_rows)]
            for pc in parsed_cells:
                if 0 <= pc.row < n_rows and 0 <= pc.col < n_cols:
                    grid[pc.row][pc.col] = pc.text

            table_id = f"p{page_no}_t{t_idx}"
            sheet_name = f"p{page_no}_t{t_idx}"[:MAX_SHEETNAME_LEN]

            tables_out.append(
                ParsedTable(
                    table_id=table_id,
                    sheet_name=sheet_name,
                    page_no=page_no,
                    table_index_on_page=t_idx,
                    bbox=t_bbox,
                    n_rows=n_rows,
                    n_cols=n_cols,
                    grid=grid,
                    merges=merges,
                    cells=parsed_cells,
                )
            )

    return tables_out, page_lines


# ============================================================
# 4) 공통 유틸/온톨로지(aliases) 기반 헤더 탐지
# ============================================================
def _norm(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", "", s)
    return s.lower()


# -----------------------------
# 페이지 유형 판별(토지 등기 페이지만 선별)
# -----------------------------
REGISTRY_MARKERS = [
    "등기사항전부증명서",
    "주요등기사항요약",
    "주요 등기사항 요약",
    "표제부",
    "갑구",
    "을구",
    "매매목록",
    "매 매 목 록",
    "부동산의 표시",
]


OTHER_DOC_MARKERS = [
    "토지이용계획확인서",
    "지적도",
    "토지 대장",
    "토지대장",
    "공유지 연명부",
    "공유지연명부",
    "건축물대장",
    "일반건축물대장",
    "집합건축물대장",
    "건축물대장 총괄표제부",
]


LAND_HEADER_MARKERS = [
    "토지[제출용]",
    "토지(제출용)",
    "토지 제출용",
]

BUILDING_HEADER_MARKERS = [
    "건물[제출용]",
    "건물(제출용)",
    "건물 제출용",
]


def compute_land_registry_pages(
    all_page_lines: Dict[int, List[PageLine]],
    total_pages: int,
    *,
    trace: Optional[List[Dict[str, Any]]] = None,
) -> List[int]:
    """
    '등기사항전부증명서 - 토지[제출용]' 시작 페이지를 기준으로
    그 다음 페이지들(연속)을 토지등기 구간으로 간주한다.

    ✅ 이유
    - 갑/을구가 이어지는 페이지는 '갑구/을구' 라벨이 없을 수 있어,
      단순 키워드 필터로는 중간 페이지가 빠지면서(예: 8-2 → 18) 순위번호가 점프함.
    - 시작 페이지의 문서 헤더(토지/건물 제출용)를 가장 신뢰하고, 상태를 carry 한다.
    """
    mode = "other"  # land | building | other
    pages: List[int] = []

    def _top_text(lines: List[PageLine], limit: int = 12) -> str:
        if not lines:
            return ""
        top = sorted(lines, key=lambda x: x.y)[:limit]
        return _norm(" ".join((ln.text or "") for ln in top))

    def _first_hit(txt: str, markers: List[str]) -> str:
        for m in markers:
            if _norm(m) in txt:
                return m
        return ""

    for p in range(1, total_pages + 1):
        prev_mode = mode
        lines = all_page_lines.get(p, [])
        txt_full = _norm(" ".join((ln.text or "") for ln in lines))
        txt_top = _top_text(lines)
        scan_txt = txt_top or txt_full
        reason = "carry"

        if not txt_full:
            # 텍스트가 거의 없는 페이지는 이전 mode를 유지 (워터마크/이하여백 등)
            include_land = mode == "land"
            if include_land:
                pages.append(p)
            reason = "empty-carry"
            if trace is not None:
                trace.append(
                    {
                        "page": p,
                        "line_count": len(lines),
                        "mode_before": prev_mode,
                        "mode_after": mode,
                        "included": include_land,
                        "reason": reason,
                    }
                )
            continue

        if "등기사항전부증명서" in scan_txt or "등기사항전부증명서" in txt_full:
            # 제출용 유형 판별
            header_txt = scan_txt if "등기사항전부증명서" in scan_txt else txt_full
            land_hit = _first_hit(header_txt, LAND_HEADER_MARKERS)
            building_hit = _first_hit(header_txt, BUILDING_HEADER_MARKERS)

            if land_hit and not building_hit:
                mode = "land"
                reason = f"header-land:{land_hit}"
            elif building_hit and not land_hit:
                mode = "building"
                reason = f"header-building:{building_hit}"
            elif ("토지" in header_txt) and ("건물" not in header_txt):
                mode = "land"
                reason = "header-land-token"
            elif ("건물" in header_txt) and ("토지" not in header_txt):
                mode = "building"
                reason = "header-building-token"
            else:
                # 토지/건물 단서가 없으면 직전 mode 유지 (보수적)
                reason = f"header-unknown-carry:{mode}"

        else:
            other_hit = _first_hit(scan_txt, OTHER_DOC_MARKERS)
            if other_hit:
                mode = "other"
                reason = f"other-doc:{other_hit}"
            else:
                reason = f"carry:{mode}"

        include_land = mode == "land"
        if include_land:
            pages.append(p)
        if trace is not None:
            trace.append(
                {
                    "page": p,
                    "line_count": len(lines),
                    "mode_before": prev_mode,
                    "mode_after": mode,
                    "included": include_land,
                    "reason": reason,
                }
            )

    # 중복 제거 + 정렬 보장
    pages = sorted(set(pages))
    return pages


def is_registry_like_page(page_lines: List[PageLine]) -> bool:
    """
    등기(토지/건물) 페이지인지 대략 판별.
    """
    if not page_lines:
        return False
    txt = _norm(" ".join((ln.text or "") for ln in page_lines))
    return any(_norm(m) in txt for m in REGISTRY_MARKERS)


def is_land_registry_page(page_lines: List[PageLine]) -> bool:
    """
    '토지 등기' 페이지인지 판별.
    - 등기 마커가 있고
    - '건물' 제출용이 아닌 페이지(건물 등기는 제외)
    - '토지' 단서가 있으면 더 신뢰, 없으면(ocr 누락)도 마커가 충분하면 통과
    """
    if not is_registry_like_page(page_lines):
        return False

    txt = _norm(" ".join((ln.text or "") for ln in page_lines))

    # 건물 등기 페이지는 제외 (제출용 헤더에 '건물'이 잘 잡히는 편)
    if "건물" in txt and "토지" not in txt:
        return False

    # 토지 단서가 있으면 OK
    if "토지" in txt or "[토지]" in txt:
        return True

    # OCR이 토지/건물 단서를 놓친 경우: 마커가 2개 이상이면 토지로 간주(보수적)
    hit_cnt = sum(1 for m in REGISTRY_MARKERS if _norm(m) in txt)
    return hit_cnt >= 2


def _contains_any(norm_text: str, aliases: List[str]) -> bool:
    for a in aliases:
        if _norm(a) and _norm(a) in norm_text:
            return True
    return False


def join_cols(row: List[str], start: int, end: int, *, sep: str = "\n") -> str:
    parts: List[str] = []
    for c in range(start, min(end, len(row))):
        v = (row[c] or "").strip()
        if v:
            parts.append(v)
    s = sep.join(parts).strip()
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s


RANK_TEXT_RE = re.compile(r"^\s*(\d+(?:-\d+)?)(?:\s*\([^)]*\))?\s*$")
RANK_PREFIX_RE = re.compile(r"^\s*(\d+(?:-\d+)?(?:\s*\([^)]*\))?)\s*(.*)$", re.S)
RANK_ANY_RE = re.compile(r"(?<!\d)(\d+(?:-\d+)?(?:\s*\([^)]*\))?)(?!\d)")


def _normalize_rank_text(v: str) -> str:
    s = (v or "").strip()
    if not s:
        return ""
    if RANK_TEXT_RE.match(s):
        return re.sub(r"\s+", " ", s)
    return ""


def _split_rank_prefix(text: str) -> Tuple[str, str]:
    """
    등기목적 등에 순위번호가 앞에 붙어서 들어온 경우 분리.
    예) '17 갑구20번 ...' -> ('17', '갑구20번 ...')
    """
    s = (text or "").strip()
    if not s:
        return "", ""
    m = RANK_PREFIX_RE.match(s)
    if not m:
        return "", s
    rank_candidate = _normalize_rank_text(m.group(1) or "")
    if not rank_candidate:
        return "", s
    rest = (m.group(2) or "").strip()
    return rank_candidate, rest


def _extract_rank_candidates(text: str) -> List[str]:
    """
    텍스트 내부의 순위번호 토큰 후보를 추출.
    (결번 복구용: '3 4', '3(전3)' 같은 케이스 보강)
    """
    s = (text or "").strip()
    if not s:
        return []
    out: List[str] = []
    for m in RANK_ANY_RE.finditer(s):
        tok = _normalize_rank_text(m.group(1) or "")
        if not tok:
            continue
        main = _rank_main_no(tok)
        # 연도/금액 등 큰 숫자 오탐 방지
        if main is None or main > 300:
            continue
        out.append(tok)
    # 순서 유지 unique
    seen = set()
    uniq = []
    for x in out:
        if x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    return uniq


SEC_DATE_RE_TEXT = r"\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일"
SEC_ACCEPTANCE_RE = re.compile(
    rf"^\s*({SEC_DATE_RE_TEXT}(?:\s*(?:제\s*)?\d+\s*호)?)(?:[\s\n]+(.*))?$",
    re.S,
)
CAUSE_ACTION_PATTERNS = [
    r"협의분할\s*에\s*의한\s*상속",
    r"전세권설정계약",
    r"근저당권설정계약",
    r"근저당권설정",
    r"저당권설정",
    r"전세권설정",
    r"지상권설정",
    r"설정계약",
    r"매매",
    r"증여",
    r"상속",
    r"해지",
    r"변경",
    r"말소",
    r"상호변경",
    r"판결",
    r"공매",
    r"경매",
    r"신탁",
    r"교환",
    r"분할",
    r"합병",
    r"가압류",
    r"압류",
    r"환매",
    r"존속기간",
]
SEC_CAUSE_ACTION_RE_TEXT = "(?:" + "|".join(CAUSE_ACTION_PATTERNS) + ")"
SEC_CAUSE_PREFIX_RE = re.compile(
    rf"^\s*({SEC_DATE_RE_TEXT}(?:\s*{SEC_CAUSE_ACTION_RE_TEXT})?)(?:[\s\n]+(.*))?$",
    re.S,
)
SEC_HOLDER_MARKER_RE = re.compile(
    r"(채권최고액|근저당권자|저당권자|전세권자|지상권자|가등기권자|권리자|공유자|소유자|채무자|목적|성명|주소|주식회사|지분)"
)
SEC_SHORT_HOLDER_LABELS = {
    "목적",
    "성명",
    "주소",
    "채무자",
    "권리자",
    "공유자",
    "소유자",
}
SEC_CAUSE_KEYWORDS = [
    "매매",
    "설정계약",
    "해지",
    "변경",
    "말소",
    "상호변경",
    "협의분할",
    "상속",
    "증여",
    "판결",
    "공매",
    "경매",
    "신탁",
    "교환",
    "분할",
    "합병",
    "가압류",
    "압류",
    "환매",
    "존속기간",
]
SEC_ADDRESS_MARKER_RE = re.compile(
    r"(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)"
)
SEC_MONEY_MARKER_RE = re.compile(r"금\s*[\d,]+원")


def _normalize_multiline_text(text: str) -> str:
    s = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+\n", "\n", s)
    s = re.sub(r"\n[ \t]+", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def _join_unique_lines(parts: List[str], *, sep: str = "\n") -> str:
    out: List[str] = []
    seen: set = set()
    for part in parts:
        txt = _normalize_multiline_text(part)
        if not txt or txt in seen:
            continue
        seen.add(txt)
        out.append(txt)
    return sep.join(out).strip()


def _append_text(base: str, extra: str, *, sep: str = "\n") -> str:
    base_norm = _normalize_multiline_text(base)
    extra_norm = _normalize_multiline_text(extra)
    if not base_norm:
        return extra_norm
    if not extra_norm:
        return base_norm
    if extra_norm in base_norm:
        return base_norm
    if base_norm in extra_norm:
        return extra_norm
    parts = base_norm.split("\n") + extra_norm.split("\n")
    return _join_unique_lines(parts, sep=sep)


def _col_interval_from_cells(
    t: ParsedTable,
    col_idx: int,
    *,
    header_row: int = -1,
) -> Optional[Tuple[float, float]]:
    def _pick_cells(require_header: bool, require_single_span: bool) -> List[ParsedCell]:
        out: List[ParsedCell] = []
        for pc in t.cells:
            if int(pc.col) != col_idx or not pc.bbox or not (pc.text or "").strip():
                continue
            if require_header and int(pc.row) != header_row:
                continue
            if require_single_span and int(pc.col_span) != 1:
                continue
            out.append(pc)
        return out

    candidates: List[ParsedCell] = []
    if header_row >= 0:
        candidates = _pick_cells(require_header=True, require_single_span=True)
        if not candidates:
            candidates = _pick_cells(require_header=True, require_single_span=False)
    if not candidates:
        candidates = _pick_cells(require_header=False, require_single_span=True)
    if not candidates:
        candidates = _pick_cells(require_header=False, require_single_span=False)
    if not candidates:
        return None

    xs0 = sorted(float(pc.bbox[0]) for pc in candidates if pc.bbox)
    xs1 = sorted(float(pc.bbox[2]) for pc in candidates if pc.bbox)
    if not xs0 or not xs1:
        return None
    mid = len(xs0) // 2
    x0 = xs0[mid]
    x1 = xs1[mid]
    if x1 <= x0:
        return None
    return (x0, x1)


def _build_sec_col_intervals(
    t: ParsedTable,
    *,
    header_row: int = -1,
    col_map: Optional[Dict[str, int]] = None,
) -> Dict[str, Tuple[float, float]]:
    cm = dict(col_map or {})
    out: Dict[str, Tuple[float, float]] = {}
    for key in ("rank", "purpose", "acceptance", "cause", "holder"):
        if key not in cm:
            continue
        try:
            col_idx = int(cm[key])
        except Exception:
            continue
        if not (0 <= col_idx < t.n_cols):
            continue
        interval = _col_interval_from_cells(t, col_idx, header_row=header_row)
        if interval is not None:
            out[key] = interval
    return out


def _best_sec_field_for_cell(
    pc: ParsedCell,
    intervals: Dict[str, Tuple[float, float]],
) -> Optional[str]:
    if not pc.bbox or not intervals:
        return None

    x0, _y0, x1, _y1 = pc.bbox
    x_mid = (float(x0) + float(x1)) / 2.0
    best_key: Optional[str] = None
    best_score = float("-inf")

    for key, (cx0, cx1) in intervals.items():
        overlap = max(0.0, min(float(x1), cx1) - max(float(x0), cx0))
        if overlap > 0:
            score = overlap
        else:
            width = max(1.0, cx1 - cx0)
            center = (cx0 + cx1) / 2.0
            score = -abs(x_mid - center) / width
        if score > best_score:
            best_score = score
            best_key = key

    return best_key


def _extract_sec_row_fields_from_cells(
    t: ParsedTable,
    row_idx: int,
    *,
    header_row: int = -1,
    col_map: Optional[Dict[str, int]] = None,
) -> Dict[str, str]:
    intervals = _build_sec_col_intervals(t, header_row=header_row, col_map=col_map)
    if not intervals:
        return {}

    row_cells = [pc for pc in t.cells if int(pc.row) == row_idx and (pc.text or "").strip()]
    if not row_cells:
        return {}

    row_cells.sort(key=lambda pc: (float(pc.bbox[0]) if pc.bbox else float(int(pc.col)), int(pc.col)))
    parts: Dict[str, List[str]] = {}
    for pc in row_cells:
        key = _best_sec_field_for_cell(pc, intervals)
        if not key:
            continue
        parts.setdefault(key, []).append(pc.text or "")

    return {key: _join_unique_lines(vals) for key, vals in parts.items() if _join_unique_lines(vals)}


def _split_acceptance_prefix(text: str) -> Tuple[str, str]:
    s = _normalize_multiline_text(text)
    if not s:
        return "", ""
    m = SEC_ACCEPTANCE_RE.match(s)
    if not m:
        return "", s
    prefix = _normalize_multiline_text(m.group(1) or "")
    rest = _normalize_multiline_text(m.group(2) or "")
    return prefix, rest


def _split_cause_prefix(text: str) -> Tuple[str, str]:
    s = _normalize_multiline_text(text)
    if not s:
        return "", ""
    m = SEC_CAUSE_PREFIX_RE.match(s)
    if not m:
        return "", s
    prefix = _normalize_multiline_text(m.group(1) or "")
    rest = _normalize_multiline_text(m.group(2) or "")
    return prefix, rest


def _is_acceptance_like(text: str) -> bool:
    return bool(SEC_ACCEPTANCE_RE.match(_normalize_multiline_text(text)))


def _is_cause_like(text: str) -> bool:
    s = _normalize_multiline_text(text)
    if not s:
        return False
    n = _norm(s)
    if re.search(SEC_DATE_RE_TEXT, s):
        return True
    return any(_norm(k) in n for k in SEC_CAUSE_KEYWORDS)


def _is_holder_like(text: str) -> bool:
    s = _normalize_multiline_text(text)
    if not s:
        return False
    if SEC_HOLDER_MARKER_RE.search(s):
        return True
    if SEC_MONEY_MARKER_RE.search(s):
        return True
    if SEC_ADDRESS_MARKER_RE.search(s):
        return True
    return ("지분" in s) or ("주식회사" in s)


def _is_short_holder_label(text: str) -> bool:
    return _normalize_multiline_text(text) in SEC_SHORT_HOLDER_LABELS


def _rebalance_sec_fields(
    *,
    purpose: str,
    acc: str,
    cause: str,
    holder: str,
) -> Tuple[str, str, str, str]:
    purpose = _normalize_multiline_text(purpose)
    acc = _normalize_multiline_text(acc)
    cause = _normalize_multiline_text(cause)
    holder = _normalize_multiline_text(holder)

    # 등기목적 칸에 접수 문구가 붙은 경우 분리
    if purpose and (not acc or not _is_acceptance_like(acc)):
        m = re.search(SEC_DATE_RE_TEXT, purpose)
        if m and m.start() > 0:
            purpose_head = _normalize_multiline_text(purpose[:m.start()])
            purpose_tail = _normalize_multiline_text(purpose[m.start():])
            tail_acc, tail_rest = _split_acceptance_prefix(purpose_tail)
            if purpose_head and tail_acc:
                purpose = purpose_head
                acc = _append_text(acc, tail_acc)
                if tail_rest:
                    cause = _append_text(cause, tail_rest)

    # 접수 칸에 등기원인 일부가 섞인 경우 분리
    acc_prefix, acc_rest = _split_acceptance_prefix(acc)
    if acc_prefix:
        acc = acc_prefix
        if acc_rest:
            cause = _append_text(acc_rest, cause)

    # 등기원인 칸 맨 앞에 접수 문구가 들어온 경우 이동
    if cause and (not acc or not _is_acceptance_like(acc)):
        moved_acc, cause_rest = _split_acceptance_prefix(cause)
        if moved_acc:
            acc = _append_text(acc, moved_acc)
            cause = cause_rest

    # 권리자 칸 맨 앞에 날짜형 등기원인이 붙은 경우 분리
    if holder:
        moved_cause, holder_rest = _split_cause_prefix(holder)
        if moved_cause and (not cause or not _is_cause_like(cause)):
            cause = _append_text(cause, moved_cause)
            holder = holder_rest

    # 등기원인 칸 안에 권리자 정보가 섞였으면 잘라서 이동
    if cause:
        marker = SEC_HOLDER_MARKER_RE.search(cause)
        if marker:
            left = _normalize_multiline_text(cause[:marker.start()])
            right = _normalize_multiline_text(cause[marker.start():])
            if right:
                if left and _is_cause_like(left):
                    cause = left
                    holder = _append_text(right, holder)
                elif not _is_cause_like(cause):
                    holder = _append_text(cause, holder)
                    cause = ""
        elif _is_holder_like(cause) and not _is_cause_like(cause):
            holder = _append_text(cause, holder)
            cause = ""

    # holder가 짧은 라벨만 남고 cause 쪽이 holder 성격이면 전부 holder로 보낸다.
    if cause and holder and _is_short_holder_label(holder) and _is_holder_like(cause) and not _is_cause_like(cause):
        holder = _append_text(cause, holder)
        cause = ""

    return (
        _normalize_multiline_text(purpose),
        _normalize_multiline_text(acc),
        _normalize_multiline_text(cause),
        _normalize_multiline_text(holder),
    )


def _extract_sec_row_fields(
    t: ParsedTable,
    row_idx: int,
    *,
    header_row: int = -1,
    col_map: Optional[Dict[str, int]] = None,
) -> Dict[str, str]:
    cm = dict(col_map or {})
    c_rank = max(0, min(int(cm.get("rank", 0)), t.n_cols - 1))
    c_purpose = max(0, min(int(cm.get("purpose", min(1, t.n_cols - 1))), t.n_cols - 1))
    c_acc = max(0, min(int(cm.get("acceptance", min(2, t.n_cols - 1))), t.n_cols - 1))
    c_cause = max(0, min(int(cm.get("cause", min(3, t.n_cols - 1))), t.n_cols - 1))
    c_holder = max(0, min(int(cm.get("holder", min(4, t.n_cols - 1))), t.n_cols - 1))

    row = t.grid[row_idx]
    fields = {
        "rank": _normalize_multiline_text(row[c_rank] or ""),
        "purpose": join_cols(row, c_purpose, c_acc, sep="\n") if c_purpose < c_acc else (row[c_purpose] or "").strip(),
        "acceptance": join_cols(row, c_acc, c_cause, sep="\n") if c_acc < c_cause else (row[c_acc] or "").strip(),
        "cause": join_cols(row, c_cause, c_holder, sep="\n") if c_cause < c_holder else (row[c_cause] or "").strip(),
        "holder": join_cols(row, c_holder, t.n_cols, sep="\n"),
    }

    cell_fields = _extract_sec_row_fields_from_cells(t, row_idx, header_row=header_row, col_map=cm)
    for key in ("rank", "purpose", "acceptance", "cause", "holder"):
        val = _normalize_multiline_text(cell_fields.get(key, ""))
        if val:
            fields[key] = val

    purpose, acc, cause, holder = _rebalance_sec_fields(
        purpose=fields.get("purpose", ""),
        acc=fields.get("acceptance", ""),
        cause=fields.get("cause", ""),
        holder=fields.get("holder", ""),
    )
    fields["purpose"] = purpose
    fields["acceptance"] = acc
    fields["cause"] = cause
    fields["holder"] = holder
    return fields


# ============================================================
# 5) 표제부 추출/정리
# ============================================================
PYO_ONTOLOGY: Dict[str, Dict[str, Any]] = {
    "display_no": {
        "label": "표시번호",
        "aliases": ["표시번호", "표시 번호", "표시no", "표시No", "표시"],
        "dtype": "string",
    },
    "acceptance": {
        "label": "접수",
        "aliases": ["접수", "접 수", "접수일자", "접수번호"],
        "dtype": "string",
    },
    "lot_address": {
        "label": "소재지번",
        "aliases": ["소재지번", "소재 지번", "소재지", "소 재 지 번", "소재지번(토지)", "소재지번(대지)"],
        "dtype": "string",
    },
    "land_category": {
        "label": "지목",
        "aliases": ["지목", "지 목"],
        "dtype": "string",
    },
    "area": {
        "label": "면적",
        "aliases": ["면적", "면 적", "면적(㎡)", "면적(m2)", "면적(m²)", "면 적(㎡)"],
        "dtype": "number",
        "unit": "㎡",
    },
}

LAND_CATEGORIES = [
    "전","답","과수원","목장용지","임야","광천지","염전","대","공장용지","학교용지","주차장","주유소용지",
    "창고용지","도로","철도용지","제방","하천","구거","유지","양어장","수도용지","공원","체육용지",
    "유원지","종교용지","사적지","묘지","잡종지"
]
LAND_CATEGORY_ALIASES = {
    "음야": "임야",
    "임아": "임야",
    "대지": "대",
}


def normalize_land_category(raw: str) -> str:
    s = (raw or "").strip()
    s2 = re.sub(r"\s+", "", s)
    if s2 in LAND_CATEGORY_ALIASES:
        s2 = LAND_CATEGORY_ALIASES[s2]
    if s2 in LAND_CATEGORIES:
        return s2
    # 약한 보정(공통 글자수)
    best = ""
    best_score = 0
    for cat in LAND_CATEGORIES:
        score = sum(1 for ch in s2 if ch in cat)
        if score > best_score:
            best_score = score
            best = cat
    if best_score >= 2 and len(s2) <= 4:
        return best
    return s


def extract_lot_no(addr: str) -> str:
    """
    소재지/주소 문자열에서 '지번' 후보를 추출.
    - 날짜(2018-01 등), 아파트 동-호(105-1103, 307-804 등) 같은 하이픈 숫자를 최대한 배제
    - '...리496-10', '...동 1429-1' 처럼 **리/동/가** 뒤에 오는 패턴을 우선시
    """
    s = (addr or "").strip()
    if not s:
        return ""

    raw_matches = list(re.finditer(r"(?<!\d)(\d{1,5}-\d{1,5})(?!\d)", s))
    candidates: List[str] = []
    ctx_matches: List[str] = []

    def is_plausible_lot(tok: str) -> bool:
        try:
            a, b = tok.split("-", 1)
            ia = int(a)
            ib = int(b)
        except Exception:
            return False

        # (1) 날짜 패턴 제거: 1900-2100 / 월 1-12
        if 1900 <= ia <= 2100 and 1 <= ib <= 12:
            return False

        # (2) 아파트 동-호(105-1103 등) 제거: 두 번째가 너무 큼
        if ib >= 1000:
            return False

        # (3) 3자리-3자리(307-804 등) 제거(지번에서 매우 드뭄)
        if len(a) == 3 and len(b) == 3:
            return False

        return True

    def is_unit_number_context(start: int, end: int) -> bool:
        left = s[max(0, start - 20):start]
        right = s[end:min(len(s), end + 12)]

        # 예: 105동 307-804호 / 2층 201-2호
        if re.search(r"\d{1,4}\s*동\s*$", left):
            return True
        if re.search(r"\d{1,3}\s*층\s*$", left):
            return True
        if re.match(r"^\s*(동|층|호|호수|호실)", right):
            return True
        if re.match(r"^\s*\d{1,4}\s*호", right):
            return True
        return False

    def has_lot_context(start: int, end: int) -> bool:
        left = s[max(0, start - 20):start]
        right = s[end:min(len(s), end + 10)]
        if re.search(r"[가-힣]{1,15}(?:리|동|가|읍|면)\s*$", left):
            return True
        if re.match(r"^\s*번지", right):
            return True
        if re.search(r"(소재지|지번)\s*$", left):
            return True
        return False

    for m in raw_matches:
        tok = m.group(1)
        st, ed = m.span(1)
        if not is_plausible_lot(tok):
            continue
        if is_unit_number_context(st, ed):
            continue
        candidates.append(tok)
        if has_lot_context(st, ed):
            ctx_matches.append(tok)

    if candidates:
        # ctx(문맥매치)가 있으면 그 중 마지막을, 없으면 전체 후보 중 마지막을 반환
        for c in reversed(ctx_matches):
            if c in candidates:
                return c
        return candidates[-1]

    # 3) 마지막 fallback: 단일 숫자(지번 문맥이 있는 경우만)
    fallback_ctx: List[str] = []
    fallback_ctx += re.findall(r"[가-힣]{1,15}(?:리|동|가|읍|면)\s*(\d{1,5})(?:\s*번지)?", s)
    fallback_ctx += re.findall(r"(?:지번|번지)\s*(\d{1,5})", s)
    if fallback_ctx:
        return fallback_ctx[-1].strip()
    return ""


def split_area_and_note(area_text: str) -> Tuple[str, str]:
    """
    면적 텍스트에서 '숫자+단위(m2/㎡/m²)'를 분리하고,
    나머지를 '등기원인 및 기타사항'으로 돌려준다.
    """
    s = (area_text or "").strip()
    if not s:
        return "", ""

    m = re.search(r"(?i)(\d+(?:,\d+)*(?:\.\d+)?)\s*(m2|m²|㎡)", s)
    if not m:
        return "", s

    num = (m.group(1) or "").replace(",", "")
    area_only = f"{num}m2"

    note = s[m.end():].strip()
    note = re.sub(r"^[\s:;,.\)\]]+", "", note).strip()
    return area_only, note



# -----------------------------
# 표제부 컬럼 추정(헤더 인식 실패/병합 헤더 대비)
# -----------------------------
_LOT_RE = re.compile(r"\b(\d{1,5}-\d{1,5})\b")


def _score_is_display_no(v: str) -> int:
    s = (v or "").strip()
    return 1 if re.fullmatch(r"\d+(?:-\d+)?", s) else 0


def _score_is_area(v: str) -> int:
    s = (v or "").strip().replace(" ", "")
    return 1 if re.search(r"(?i)\d+(?:,\d+)*(?:\.\d+)?(m2|m²|㎡)", s) else 0


def _score_is_land_category(v: str) -> int:
    s = normalize_land_category(v)
    return 1 if s in LAND_CATEGORIES else 0


def _score_is_acceptance(v: str) -> int:
    s = (v or "").strip()
    # 날짜(YYYY년M월D일) 또는 접수번호(제xxxx호) 패턴
    if re.search(r"\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일", s):
        return 2
    if re.search(r"제\s*\d+\s*호", s):
        return 1
    return 0


def _score_is_address(v: str) -> int:
    s = (v or "").strip()
    score = 0
    if _LOT_RE.search(s):
        score += 2
    if re.search(r"(시|군|구|읍|면|동|리)", s):
        score += 1
    if "소재지" in s or "지번" in s:
        score += 1
    return score


def infer_pyo_col_map(t: ParsedTable, header_row: int) -> Dict[str, int]:
    """
    헤더가 병합/오독으로 컬럼명이 안 보이는 경우를 위해,
    데이터 행 패턴으로 표제부 컬럼을 추정한다.
    """
    # 샘플링 범위
    r0 = header_row + 1
    r1 = min(t.n_rows, header_row + 1 + 30)

    # 컬럼별 점수 계산
    scores = []
    for c in range(t.n_cols):
        disp = area = cat = acc = addr = 0
        for r in range(r0, r1):
            v = (t.grid[r][c] or "").strip()
            if not v:
                continue
            disp += _score_is_display_no(v)
            area += _score_is_area(v)
            cat += _score_is_land_category(v)
            acc += _score_is_acceptance(v)
            addr += _score_is_address(v)
        scores.append({"c": c, "disp": disp, "area": area, "cat": cat, "acc": acc, "addr": addr})

    # best columns
    c_disp = max(scores, key=lambda x: x["disp"])["c"] if scores else 0
    c_area = max(scores, key=lambda x: x["area"])["c"] if scores else min(3, t.n_cols - 1)
    c_cat = max(scores, key=lambda x: x["cat"])["c"] if scores else max(0, min(c_area - 1, t.n_cols - 1))

    # address는 지목 앞쪽에서 찾는 게 안전
    cand_addr = [s for s in scores if s["c"] < c_cat]
    c_addr = max(cand_addr, key=lambda x: x["addr"])["c"] if cand_addr else max(0, min(c_cat - 1, t.n_cols - 1))

    # acceptance는 표시번호 다음~소재지번 이전에서 찾는 게 안전
    cand_acc = [s for s in scores if c_disp < s["c"] < c_addr]
    if cand_acc:
        c_acc = max(cand_acc, key=lambda x: x["acc"])["c"]
    else:
        c_acc = min(c_disp + 1, t.n_cols - 1)

    col_map = {
        "display_no": c_disp,
        "acceptance": c_acc,
        "lot_address": c_addr,
        "land_category": c_cat,
        "area": c_area,
    }

    # 순서가 뒤집힌 경우 보정(최소한 display_no < land_category < area)
    if not (col_map["display_no"] < col_map["land_category"] < col_map["area"]):
        # land_category/area가 바뀐 경우가 많음 → swap 시도
        if col_map["display_no"] < col_map["area"] < col_map["land_category"]:
            col_map["land_category"], col_map["area"] = col_map["area"], col_map["land_category"]

    return col_map
def find_pyo_tables(tables: List[ParsedTable]) -> List[Tuple[ParsedTable, int, Dict[str, int]]]:
    """
    표제부 테이블 후보 탐색.
    - 헤더가 명확하면 헤더 기반 매핑
    - 헤더가 병합/오독이면 데이터 패턴으로 컬럼 추정(infer_pyo_col_map)
    """
    out: List[Tuple[ParsedTable, int, Dict[str, int]]] = []

    for t in tables:
        if t.n_rows < 2 or t.n_cols < 3:
            continue

        header_row = -1
        for r in range(min(t.n_rows, 15)):  # ✅ 탐색 범위 확대
            row_norm = _norm(" ".join(t.grid[r]))
            hits = {k: 1 for k, spec in PYO_ONTOLOGY.items() if _contains_any(row_norm, spec["aliases"])}

            # 케이스1) 표시번호 포함 + 최소 3개 필드 히트
            cond1 = ("display_no" in hits) and (len(hits) >= 3)
            # 케이스2) 표시번호가 안 보여도 소재지번/지목/면적 3종이 보이면 표제부로 간주
            cond2 = ("lot_address" in hits) and ("land_category" in hits) and ("area" in hits)
            # 케이스3) '표제부' 텍스트가 섞여 있고 지목/면적이 보이면 표제부 가능성
            cond3 = ("표제부" in row_norm) and (("land_category" in hits) and ("area" in hits))

            if cond1 or cond2 or cond3:
                header_row = r
                break

        if header_row < 0:
            continue

        # 1차: 헤더 기반 매핑
        col_map: Dict[str, int] = {}
        for c in range(t.n_cols):
            cell_norm = _norm(t.grid[header_row][c])

            if "display_no" not in col_map and _contains_any(cell_norm, PYO_ONTOLOGY["display_no"]["aliases"]):
                col_map["display_no"] = c
                continue
            if "acceptance" not in col_map and _contains_any(cell_norm, PYO_ONTOLOGY["acceptance"]["aliases"]):
                col_map["acceptance"] = c
                continue
            if "lot_address" not in col_map and _contains_any(cell_norm, PYO_ONTOLOGY["lot_address"]["aliases"]):
                col_map["lot_address"] = c
                continue
            if "land_category" not in col_map and _contains_any(cell_norm, PYO_ONTOLOGY["land_category"]["aliases"]):
                col_map["land_category"] = c
                continue
            if "area" not in col_map and _contains_any(cell_norm, PYO_ONTOLOGY["area"]["aliases"]):
                col_map["area"] = c
                continue

        # 2차: 부족한 컬럼은 데이터 패턴으로 추정
        need_infer = ("land_category" not in col_map) or ("area" not in col_map) or ("lot_address" not in col_map)
        if need_infer:
            inferred = infer_pyo_col_map(t, header_row)
            for k, v in inferred.items():
                col_map.setdefault(k, v)

        # acceptance가 없으면 표시번호 다음칸으로 보수적 추정
        col_map.setdefault("display_no", 0)
        if "acceptance" not in col_map:
            cand = col_map.get("display_no", 0) + 1
            if cand < t.n_cols:
                col_map["acceptance"] = cand

        # sanity: 표시번호 < 지목 < 면적 (일반적)
        if not (col_map["display_no"] < col_map["land_category"] < col_map["area"]):
            # 마지막으로 infer 재시도(간혹 헤더 매핑이 뒤틀림)
            col_map = infer_pyo_col_map(t, header_row)
            if not (col_map["display_no"] < col_map["land_category"] < col_map["area"]):
                continue

        out.append((t, header_row, col_map))

    return out


def extract_pyo_records_from_table(t: ParsedTable, header_row: int, col_map: Dict[str, int]) -> pd.DataFrame:
    c_disp = col_map["display_no"]
    c_acc = col_map.get("acceptance")
    c_addr = col_map["lot_address"]
    c_cat = col_map["land_category"]
    c_area = col_map["area"]

    records: List[Dict[str, Any]] = []
    cur: Optional[Dict[str, Any]] = None

    for r in range(header_row + 1, t.n_rows):
        row = t.grid[r]
        if all((not (row[c] or "").strip()) for c in range(t.n_cols)):
            continue

        disp = (row[c_disp] or "").strip()
        # 헤더/잡음 행 skip
        if disp and _contains_any(_norm(disp), ["표제부", "갑구", "을구", "순위번호"]):
            continue

        # 접수: 접수 col ~ 소재지번 직전
        acc_text = ""
        if c_acc is not None and 0 <= c_acc < t.n_cols:
            if c_acc < c_addr:
                acc_text = join_cols(row, c_acc, c_addr, sep="\n")
            else:
                acc_text = (row[c_acc] or "").strip()

        # 소재지번: 소재지번 col ~ 지목 직전
        lot_addr = join_cols(row, c_addr, c_cat, sep="\n")

        # 지목: 지목 col ~ 면적 직전
        land_cat_raw = join_cols(row, c_cat, c_area, sep=" ")

        # 면적: 면적 col ~ 끝
        area_text_raw = join_cols(row, c_area, t.n_cols, sep=" ")
        area_only, note_text = split_area_and_note(area_text_raw)

        is_cont = (not disp) and (acc_text or lot_addr or land_cat_raw or area_only or note_text)
        if is_cont and cur is not None:
            if acc_text:
                cur["접수"] = (cur["접수"] + "\n" + acc_text).strip() if cur["접수"] else acc_text
            if lot_addr:
                cur["소재지번"] = (cur["소재지번"] + "\n" + lot_addr).strip() if cur["소재지번"] else lot_addr
            if land_cat_raw:
                cur["_지목_raw"] = (cur["_지목_raw"] + " " + land_cat_raw).strip() if cur["_지목_raw"] else land_cat_raw
            if area_only and not cur.get("면적"):
                cur["면적"] = area_only
            if note_text:
                cur["등기원인 및 기타사항"] = (
                    (cur.get("등기원인 및 기타사항") or "").strip() + "\n" + note_text
                ).strip() if (cur.get("등기원인 및 기타사항") or "").strip() else note_text
            continue

        if disp:
            cur = {
                "페이지": t.page_no,
                "table_id": t.table_id,
                "표시번호": disp,
                "접수": acc_text,
                "소재지번": lot_addr,
                "_지목_raw": land_cat_raw,
                "면적": area_only,
                "등기원인 및 기타사항": note_text,
            }
            records.append(cur)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    df["지번"] = df["소재지번"].apply(extract_lot_no)
    df["지목"] = df["_지목_raw"].apply(normalize_land_category)
    df = df.drop(columns=["_지목_raw"], errors="ignore")

    # 정렬(표시번호)
    def disp_key(x: Any) -> Tuple[int, int, str]:
        s = str(x or "").strip()
        m = re.match(r"^(\d+)(?:-(\d+))?", s)
        if not m:
            return (10**9, 0, s)
        a = int(m.group(1))
        b = int(m.group(2) or 0)
        return (a, b, s)

    df["_k1"] = df["표시번호"].apply(disp_key)
    df = df.sort_values("_k1").drop(columns=["_k1"])

    cols = ["페이지", "table_id", "표시번호", "접수", "소재지번", "지번", "지목", "면적", "등기원인 및 기타사항"]
    return df[[c for c in cols if c in df.columns]]


def extract_eul_records_from_table(t: ParsedTable, header_row: int, col_map: Dict[str, int]) -> pd.DataFrame:
    """
    을구(소유권 이외의 권리) 테이블에서 레코드 추출.
    구조는 갑구와 동일하지만, 갑구 성격 행(소유권/압류 등)이 섞이면 제거한다.
    """
    c_rank = col_map["rank"]
    c_purpose = col_map["purpose"]
    c_acc = col_map["acceptance"]
    c_cause = col_map["cause"]
    c_holder = col_map["holder"]

    records: List[Dict[str, Any]] = []
    cur: Optional[Dict[str, Any]] = None

    for r in range(header_row + 1, t.n_rows):
        row = t.grid[r]
        if all((not (row[c] or "").strip()) for c in range(t.n_cols)):
            continue

        row_fields = _extract_sec_row_fields(t, r, header_row=header_row, col_map=col_map)
        rank = _normalize_rank_text(row_fields.get("rank", ""))

        # 잡음/헤더 반복 제거
        if rank and _contains_any(_norm(rank), ["순위번호", "갑구", "을구", "표제부"]):
            continue

        purpose = row_fields.get("purpose", "")
        acc = row_fields.get("acceptance", "")
        cause = row_fields.get("cause", "")
        holder = row_fields.get("holder", "")

        # 순위번호가 칸 어긋남으로 다른 열/등기목적에 들어온 경우 복구
        if not rank:
            scan_cols = min(max(c_rank + 2, 3), t.n_cols)
            for c in range(scan_cols):
                if c == c_rank:
                    continue
                probe = _normalize_rank_text(row[c] or "")
                if probe:
                    rank = probe
                    break
        if not rank and purpose:
            rank_from_purpose, purpose_rest = _split_rank_prefix(purpose)
            if rank_from_purpose:
                rank = rank_from_purpose
                purpose = purpose_rest or purpose

        purpose, acc, cause, holder = _rebalance_sec_fields(
            purpose=purpose,
            acc=acc,
            cause=cause,
            holder=holder,
        )

        # continuation: 순위번호가 비어있고 내용이 있으면 이전 레코드에 병합
        is_cont = (not rank) and (purpose or acc or cause or holder)
        if is_cont and cur is not None:
            if purpose:
                cur["등기목적"] = (cur["등기목적"] + "\n" + purpose).strip() if cur["등기목적"] else purpose
            if acc:
                cur["접수"] = (cur["접수"] + "\n" + acc).strip() if cur["접수"] else acc
            if cause:
                cur["등기원인"] = (cur["등기원인"] + "\n" + cause).strip() if cur["등기원인"] else cause
            if holder:
                cur["권리자 및 기타사항"] = (
                    (cur.get("권리자 및 기타사항") or "").strip() + "\n" + holder
                ).strip() if (cur.get("권리자 및 기타사항") or "").strip() else holder
            continue

        if rank:
            cur = {
                "페이지": t.page_no,
                "table_id": t.table_id,
                "순위번호": rank,
                "등기목적": purpose,
                "접수": acc,
                "등기원인": cause,
                "권리자 및 기타사항": holder,
            }
            records.append(cur)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # ✅ 갑구 성격의 행이 섞이면 을구 결과가 깨지므로 제거
    if "등기목적" in df.columns:
        df["_ptype"] = df["등기목적"].apply(_purpose_type)
        df = df[df["_ptype"] != "gab"].drop(columns=["_ptype"], errors="ignore")
        if df.empty:
            return df

    # 정렬(순위번호)
    def rank_key(x: Any) -> Tuple[int, int, str]:
        s = str(x or "").strip()
        m = re.match(r"^(\d+)(?:-(\d+))?", s)
        if not m:
            return (10**9, 0, s)
        a = int(m.group(1))
        b = int(m.group(2) or 0)
        return (a, b, s)

    df["_rk"] = df["순위번호"].apply(rank_key)
    df = df.sort_values("_rk").drop(columns=["_rk"])

    cols = ["페이지", "table_id", "순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
    return df[[c for c in cols if c in df.columns]]



# ============================================================
# 6) 갑구 추출/정리
# ============================================================
GAB_ONTOLOGY: Dict[str, Dict[str, Any]] = {
    "rank": {
        "label": "순위번호",
        "aliases": ["순위번호", "순위 번호", "순 위 번 호", "순위", "순위No", "순위NO"],
    },
    "purpose": {
        "label": "등기목적",
        "aliases": ["등기목적", "등기 목적", "등 기 목 적"],
    },
    "acceptance": {
        "label": "접수",
        "aliases": ["접수", "접 수", "접수일자", "접수번호"],
    },
    "cause": {
        "label": "등기원인",
        "aliases": ["등기원인", "등기 원인", "등 기 원 인"],
    },
    "holder": {
        "label": "권리자 및 기타사항",
        "aliases": ["권리자및기타사항", "권리자 및 기타사항", "권리자및 기타사항", "권리자 및기타사항", "권리자 및 기타 사항"],
    },
}


def classify_gab_or_eul(table: ParsedTable, page_lines: List[PageLine]) -> str:
    """
    테이블이 '갑구/을구' 중 어디에 속하는지 분류.
    1) page_lines(일반 OCR 텍스트)에서 테이블 바로 위의 '갑 구/을 구' 라벨을 찾는다.
    2) 없으면, 테이블 상단 몇 행(grid)에서 라벨을 찾는다.
    3) 그래도 없으면 unknown.
    """
    # table y0 추정 (bbox가 None이면 cells bbox로 다시 계산)
    y0 = None
    if table.bbox:
        y0 = table.bbox[1]
    else:
        ys = []
        for pc in table.cells:
            if pc.bbox:
                ys.append(pc.bbox[1])
        if ys:
            y0 = min(ys)

    def is_gab_marker(txt: str) -> bool:
        n = _norm(txt)
        # '갑구' 직접 또는 '소유권에 관한 사항' (이외/권리 제외)로 판정
        if "갑구" in n:
            return True
        if "소유권에관한사항" in n and "이외" not in n:
            return True
        # OCR 흔들림: '값구','각구' 등
        if ("값구" in n) or ("각구" in n):
            return True
        return False

    def is_eul_marker(txt: str) -> bool:
        n = _norm(txt)
        if "을구" in n:
            return True
        if "소유권이외의권리에관한사항" in n:
            return True
        # OCR 흔들림
        if "을구" in n:
            return True
        return False

    # 1) page_lines 기준: 테이블 위쪽에서 가장 가까운 마커
    if y0 is not None and page_lines:
        cand = [ln for ln in page_lines if ln.y < y0]
        # 가까운 것부터 역순
        cand.sort(key=lambda ln: ln.y, reverse=True)

        # 우선 가까운 범위(<=800px)에서 찾고, 없으면 전체에서 찾기
        for max_gap in (800, 2000, 10**9):
            for ln in cand:
                if (y0 - ln.y) > max_gap:
                    break
                if is_gab_marker(ln.text):
                    return "갑구"
                if is_eul_marker(ln.text):
                    return "을구"

    # 2) table grid 상단에서 라벨 탐색 (가끔 라벨이 테이블 내부로 들어오는 경우)
    top_rows = min(5, table.n_rows)
    for r in range(top_rows):
        row_text = " ".join(table.grid[r])
        if is_gab_marker(row_text):
            return "갑구"
        if is_eul_marker(row_text):
            return "을구"

    return "unknown"


def _looks_like_sec_continuation_table(t: ParsedTable) -> bool:
    """
    헤더가 누락/파손된 갑·을구 연속 테이블을 약식 탐지.
    조건:
    - 행/열이 너무 작지 않고
    - 상단 행들에서 순위번호 형태가 2회 이상 보임
    - 표제부 고유 헤더(소재지번/지목/면적)가 강하게 보이지 않음
    """
    meta = _analyze_sec_table_candidate(t)
    return bool(meta.get("looks_like_continuation"))


def _analyze_sec_table_candidate(t: ParsedTable) -> Dict[str, Any]:
    """
    갑/을구 연속표 가능성을 디버그 가능한 형태로 계산한다.
    """
    out: Dict[str, Any] = {
        "page": int(t.page_no),
        "table_id": str(t.table_id),
        "n_rows": int(t.n_rows),
        "n_cols": int(t.n_cols),
        "header_row": -1,
        "header_hits": [],
        "nonempty_rows": 0,
        "rank_like_rows": 0,
        "purpose_keyword_rows": 0,
        "date_like_rows": 0,
        "holder_like_rows": 0,
        "multi_cell_rows": 0,
        "continuation_score": 0,
        "pyo_like": False,
        "looks_like_continuation": False,
    }

    if t.n_rows < 2 or t.n_cols < 3:
        return out

    head_txt = _norm(" ".join(" ".join(t.grid[r]) for r in range(min(t.n_rows, 3))))
    out["pyo_like"] = any(k in head_txt for k in map(_norm, ["소재지번", "지목", "면적", "표시번호"]))

    for r in range(min(t.n_rows, 10)):
        row_norm = _norm(" ".join(t.grid[r]))
        hits = [k for k, spec in GAB_ONTOLOGY.items() if _contains_any(row_norm, spec["aliases"])]
        if ("rank" in hits) and (
            len(hits) >= 4
            or (("purpose" in hits) and ("acceptance" in hits))
            or (("purpose" in hits) and ("holder" in hits))
        ):
            out["header_row"] = int(r)
            out["header_hits"] = hits
            break

    for r in range(min(t.n_rows, 30)):
        row = t.grid[r]
        cells = [(x or "").strip() for x in row]
        nonempty = [x for x in cells if x]
        if not nonempty:
            continue

        out["nonempty_rows"] += 1
        if len(nonempty) >= min(3, max(2, t.n_cols)):
            out["multi_cell_rows"] += 1

        left_cells = cells[: min(3, t.n_cols)]
        has_rank = False
        for cell in left_cells:
            if _normalize_rank_text(cell):
                has_rank = True
                break
            if _split_rank_prefix(cell)[0]:
                has_rank = True
                break
            if _extract_rank_candidates(cell):
                has_rank = True
                break
        if has_rank:
            out["rank_like_rows"] += 1

        row_text = " ".join(nonempty)
        purpose_probe = " ".join(cells[1:min(t.n_cols, 4)]).strip() if t.n_cols > 1 else row_text
        if _purpose_type(purpose_probe or row_text) in ("gab", "eul"):
            out["purpose_keyword_rows"] += 1

        if re.search(r"\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일", row_text) or re.search(r"제\s*\d+\s*호", row_text):
            out["date_like_rows"] += 1

        if t.n_cols >= 4:
            if any(cells[c] for c in range(min(4, t.n_cols - 1), t.n_cols)):
                out["holder_like_rows"] += 1
        elif len(nonempty) >= 3:
            out["holder_like_rows"] += 1

    score = 0
    if out["rank_like_rows"] >= 2:
        score += 2
    elif out["rank_like_rows"] >= 1:
        score += 1
    if out["purpose_keyword_rows"] >= 2:
        score += 2
    elif out["purpose_keyword_rows"] >= 1:
        score += 1
    if out["date_like_rows"] >= 2:
        score += 2
    elif out["date_like_rows"] >= 1:
        score += 1
    if out["holder_like_rows"] >= 2:
        score += 1
    if out["multi_cell_rows"] >= 2:
        score += 1

    out["continuation_score"] = score
    out["looks_like_continuation"] = (not out["pyo_like"]) and out["nonempty_rows"] >= 2 and (
        out["rank_like_rows"] >= 2
        or score >= 4
        or (
            out["rank_like_rows"] >= 1
            and out["purpose_keyword_rows"] >= 1
            and out["date_like_rows"] >= 1
        )
    )
    return out


def find_gab_tables(
    tables: List[ParsedTable],
) -> List[Tuple[ParsedTable, int, Dict[str, int]]]:
    """
    '순위번호/등기목적/접수/등기원인/권리자및기타사항' 헤더가 있는 테이블 후보 탐색
    (갑구/을구 공통 구조라서, 실제 갑/을 구 판별은 page_lines + bbox로 추가 분류)
    """
    out: List[Tuple[ParsedTable, int, Dict[str, int]]] = []

    for t in tables:
        if t.n_rows < 2 or t.n_cols < 3:
            continue

        header_row = -1
        for r in range(min(t.n_rows, 10)):
            row_norm = _norm(" ".join(t.grid[r]))
            hits = {k: 1 for k, spec in GAB_ONTOLOGY.items() if _contains_any(row_norm, spec["aliases"])}
            # 최소 조건: 순위번호 + 권리자 + 접수 + 등기목적 중 3~4개
            if ("rank" in hits) and (
                len(hits) >= 4
                or (("purpose" in hits) and ("acceptance" in hits))
                or (("purpose" in hits) and ("holder" in hits))
            ):
                header_row = r
                break

        col_map: Dict[str, int] = {}
        if header_row >= 0:
            for c in range(t.n_cols):
                cell_norm = _norm(t.grid[header_row][c])

                if "rank" not in col_map and _contains_any(cell_norm, GAB_ONTOLOGY["rank"]["aliases"]):
                    col_map["rank"] = c
                    continue
                if "purpose" not in col_map and _contains_any(cell_norm, GAB_ONTOLOGY["purpose"]["aliases"]):
                    col_map["purpose"] = c
                    continue
                if "acceptance" not in col_map and _contains_any(cell_norm, GAB_ONTOLOGY["acceptance"]["aliases"]):
                    col_map["acceptance"] = c
                    continue
                if "cause" not in col_map and _contains_any(cell_norm, GAB_ONTOLOGY["cause"]["aliases"]):
                    col_map["cause"] = c
                    continue
                if "holder" not in col_map and _contains_any(cell_norm, GAB_ONTOLOGY["holder"]["aliases"]):
                    col_map["holder"] = c
                    continue
        else:
            # 헤더를 못 찾았지만 순위번호 패턴이 충분하면 연속표 후보로 채택
            if not _looks_like_sec_continuation_table(t):
                continue

        # fallback (일반적인 컬럼 순서)
        col_map.setdefault("rank", 0)
        col_map.setdefault("purpose", 1)
        col_map.setdefault("acceptance", min(2, t.n_cols - 1))
        col_map.setdefault("cause", min(3, t.n_cols - 1))
        col_map.setdefault("holder", min(4, t.n_cols - 1))

        # 순서 보정: rank < purpose < acceptance < cause < holder
        # 만약 뒤죽박죽이면 오탐일 가능성이 높아서 제외
        try:
            if not (col_map["rank"] <= col_map["purpose"] <= col_map["acceptance"] <= col_map["cause"] <= col_map["holder"]):
                # 그래도 holder가 맨 끝에 오도록 강제
                col_map["holder"] = max(col_map["holder"], col_map["cause"], col_map["acceptance"], col_map["purpose"])
                if col_map["holder"] >= t.n_cols:
                    col_map["holder"] = t.n_cols - 1
        except Exception:
            continue

        out.append((t, header_row, col_map))

    return out


GAB_PURPOSE_KEYWORDS = [
    # 갑구(소유권/압류/가처분 등) - 강한 키워드 중심
    "소유권", "공유", "압류", "가압류", "가처분", "경매", "환매", "가등기", "신탁", "처분금지", "보존등기",
]
GAB_WEAK_PURPOSE_KEYWORDS = [
    # '변경/이전/말소'는 을구에도 많이 나오므로 약한 신호로만 취급
    "말소", "이전", "변경", "회복", "보존",
]
EUL_PURPOSE_KEYWORDS = [
    # 을구(담보/임차권 등)
    "근저당", "저당", "전세권", "지상권", "임차권", "지역권", "담보", "질권", "저당권", "근질권",
]


def _purpose_type(purpose_text: str) -> str:
    """
    등기목적 텍스트 기반으로 갑구/을구 성격을 판정.
    return: 'gab' | 'eul' | 'unknown'
    """
    n = _norm(purpose_text or "")
    gab_strong = any(k in n for k in map(_norm, GAB_PURPOSE_KEYWORDS))
    gab_weak = any(k in n for k in map(_norm, GAB_WEAK_PURPOSE_KEYWORDS))
    eul_strong = any(k in n for k in map(_norm, EUL_PURPOSE_KEYWORDS))

    # 을구 강신호가 있으면 우선적으로 을구로 본다.
    # (예: 근저당권이전/변경/말소는 을구에서 매우 흔함)
    if eul_strong and not gab_strong:
        return "eul"

    if gab_strong and not eul_strong:
        return "gab"

    if gab_strong and eul_strong:
        # 둘 다 강신호면 소유권 단어가 있으면 갑구, 아니면 을구 우선
        if "소유권" in n and "이외" not in n:
            return "gab"
        return "eul"

    # 약신호만 있을 때는 unknown (행 제거 방지)
    if gab_weak:
        return "unknown"
    return "unknown"


def guess_section_by_purpose_keywords(t: ParsedTable, header_row: int, col_map: Dict[str, int]) -> str:
    """
    '등기목적' 텍스트를 보고 갑구/을구를 추정하는 백업 로직.
    - 소유권/압류/가처분 등 → 갑구 가능성↑
    - 근저당/전세권/지상권/임차권 등 → 을구 가능성↑
    """
    c_purpose = col_map.get("purpose", 1)
    c_next = col_map.get("acceptance", min(c_purpose + 1, t.n_cols))
    gab = 0
    eul = 0

    for r in range(header_row + 1, min(t.n_rows, header_row + 30)):
        txt = join_cols(t.grid[r], c_purpose, c_next, sep=" ")
        if not _norm(txt):
            continue
        ptype = _purpose_type(txt)
        if ptype == "gab":
            gab += 1
        elif ptype == "eul":
            eul += 1

    if gab == 0 and eul == 0:
        return "unknown"
    if gab == eul:
        return "unknown"
    return "갑구" if gab > eul else "을구"
def extract_gab_records_from_table(t: ParsedTable, header_row: int, col_map: Dict[str, int]) -> pd.DataFrame:
    c_rank = col_map["rank"]
    c_purpose = col_map["purpose"]
    c_acc = col_map["acceptance"]
    c_cause = col_map["cause"]
    c_holder = col_map["holder"]

    records: List[Dict[str, Any]] = []
    cur: Optional[Dict[str, Any]] = None

    for r in range(header_row + 1, t.n_rows):
        row = t.grid[r]
        if all((not (row[c] or "").strip()) for c in range(t.n_cols)):
            continue

        row_fields = _extract_sec_row_fields(t, r, header_row=header_row, col_map=col_map)
        rank = _normalize_rank_text(row_fields.get("rank", ""))

        # 잡음/헤더 반복 제거
        if rank and _contains_any(_norm(rank), ["순위번호", "갑구", "을구", "표제부"]):
            continue

        purpose = row_fields.get("purpose", "")
        acc = row_fields.get("acceptance", "")
        cause = row_fields.get("cause", "")
        holder = row_fields.get("holder", "")

        # 순위번호가 칸 어긋남으로 다른 열/등기목적에 들어온 경우 복구
        if not rank:
            scan_cols = min(max(c_rank + 2, 3), t.n_cols)
            for c in range(scan_cols):
                if c == c_rank:
                    continue
                probe = _normalize_rank_text(row[c] or "")
                if probe:
                    rank = probe
                    break
        if not rank and purpose:
            rank_from_purpose, purpose_rest = _split_rank_prefix(purpose)
            if rank_from_purpose:
                rank = rank_from_purpose
                purpose = purpose_rest or purpose

        purpose, acc, cause, holder = _rebalance_sec_fields(
            purpose=purpose,
            acc=acc,
            cause=cause,
            holder=holder,
        )

        # continuation: 순위번호가 비어있고 내용이 있으면 이전 레코드에 병합
        is_cont = (not rank) and (purpose or acc or cause or holder)
        if is_cont and cur is not None:
            if purpose:
                cur["등기목적"] = (cur["등기목적"] + "\n" + purpose).strip() if cur["등기목적"] else purpose
            if acc:
                cur["접수"] = (cur["접수"] + "\n" + acc).strip() if cur["접수"] else acc
            if cause:
                cur["등기원인"] = (cur["등기원인"] + "\n" + cause).strip() if cur["등기원인"] else cause
            if holder:
                cur["권리자 및 기타사항"] = (
                    (cur.get("권리자 및 기타사항") or "").strip() + "\n" + holder
                ).strip() if (cur.get("권리자 및 기타사항") or "").strip() else holder
            continue

        if rank:
            cur = {
                "페이지": t.page_no,
                "table_id": t.table_id,
                "순위번호": rank,
                "등기목적": purpose,
                "접수": acc,
                "등기원인": cause,
                "권리자 및 기타사항": holder,
            }
            records.append(cur)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # ✅ 을구(근저당/전세권/지상권 등) 성격의 행이 섞이면 갑구 결과가 깨지므로 제거
    if "등기목적" in df.columns:
        df["_ptype"] = df["등기목적"].apply(_purpose_type)
        df = df[df["_ptype"] != "eul"].drop(columns=["_ptype"], errors="ignore")
        if df.empty:
            return df

    # 정렬(순위번호)
    def rank_key(x: Any) -> Tuple[int, int, str]:
        s = str(x or "").strip()
        m = re.match(r"^(\d+)(?:-(\d+))?", s)
        if not m:
            return (10**9, 0, s)
        a = int(m.group(1))
        b = int(m.group(2) or 0)
        return (a, b, s)

    df["_rk"] = df["순위번호"].apply(rank_key)
    df = df.sort_values("_rk").drop(columns=["_rk"])

    cols = ["페이지", "table_id", "순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
    return df[[c for c in cols if c in df.columns]]


# ============================================================
# 7) 지번(필지)별 그룹핑
# ============================================================
@dataclass
class ParcelGroup:
    key: str
    start_page: int
    end_page: int
    pyo_tables: List[ParsedTable] = field(default_factory=list)
    pyo_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    gab_tables: List[ParsedTable] = field(default_factory=list)
    gab_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    eul_tables: List[ParsedTable] = field(default_factory=list)
    eul_df: pd.DataFrame = field(default_factory=pd.DataFrame)


def _pick_group_key_from_pyo_df(df: pd.DataFrame, fallback: str) -> str:
    if "지번" in df.columns:
        vals = [str(v).strip() for v in df["지번"].tolist() if str(v).strip()]
        if vals:
            return vals[0]
    return fallback


def group_parcels_from_pyo(
    pyo_items: List[Tuple[ParsedTable, pd.DataFrame]],
    total_pages: int,
) -> List[ParcelGroup]:
    """
    표제부 결과를 기반으로 지번별 그룹을 만들고, 페이지 범위를 추정한다.
    - start_page: 해당 지번의 표제부가 처음 등장한 페이지
    - end_page: 다음 지번 start_page - 1 (마지막은 total_pages)
    """
    tmp: Dict[str, Dict[str, Any]] = {}
    for t, df in pyo_items:
        key = _pick_group_key_from_pyo_df(df, fallback=t.table_id)
        g = tmp.setdefault(key, {"start": t.page_no, "tables": [], "dfs": []})
        g["start"] = min(g["start"], t.page_no)
        g["tables"].append(t)
        g["dfs"].append(df)

    groups: List[ParcelGroup] = []
    for key, g in tmp.items():
        pyo_df = pd.concat(g["dfs"], ignore_index=True) if g["dfs"] else pd.DataFrame()
        pyo_df = pyo_df.drop_duplicates()
        groups.append(ParcelGroup(key=key, start_page=int(g["start"]), end_page=total_pages, pyo_tables=g["tables"], pyo_df=pyo_df))

    groups.sort(key=lambda x: x.start_page)

    # end_page 계산
    for i in range(len(groups)):
        if i < len(groups) - 1:
            groups[i].end_page = max(groups[i].start_page, groups[i + 1].start_page - 1)
        else:
            groups[i].end_page = total_pages

    return groups


def assign_table_to_group(groups: List[ParcelGroup], page_no: int) -> Optional[ParcelGroup]:
    for g in groups:
        if g.start_page <= page_no <= g.end_page:
            return g
    return None


def _normalize_lot_key(v: Any) -> str:
    return re.sub(r"\s+", "", str(v or "").strip())


def _is_probable_lot_key(v: Any) -> bool:
    return re.fullmatch(r"\d{1,5}(?:-\d{1,5})?", _normalize_lot_key(v)) is not None


def find_group_by_lot_key(groups: List[ParcelGroup], lot_key: str) -> Optional[ParcelGroup]:
    nk = _normalize_lot_key(lot_key)
    if not _is_probable_lot_key(nk):
        return None
    for g in groups:
        if _is_probable_lot_key(g.key) and _normalize_lot_key(g.key) == nk:
            return g
    return None


# -----------------------------
# (보강) 갑구 테이블을 지번(필지) 그룹에 더 정확히 매핑하기 위한 lot 추정
# -----------------------------
_LOT_SHORT_RE = re.compile(r"\b(\d{1,5}-\d{1,5})\b")


def guess_lot_from_lines(page_lines: List[PageLine]) -> str:
    """
    페이지의 텍스트 라인에서 지번 후보를 추정.
    - 1~5자리-1~5자리 패턴만 사용(주민/법인번호 등 긴 패턴 배제)
    - '리/동/지번/소재지' 같은 문맥이 있으면 가중치
    """
    if not page_lines:
        return ""

    scores: Dict[str, float] = {}
    # 상단부에 보통 소재지가 있으므로 위쪽(작은 y)부터 30줄만
    lines_sorted = sorted(page_lines, key=lambda x: x.y)[:30]
    for ln in lines_sorted:
        txt = ln.text or ""
        for lot in _LOT_SHORT_RE.findall(txt):
            sc = 1.0
            if re.search(r"(리|동)\s*" + re.escape(lot), txt):
                sc += 2.0
            if ("소재지" in txt) or ("지번" in txt) or ("토지" in txt):
                sc += 1.0
            # 위쪽일수록 가중치(0~1)
            sc += max(0.0, 1.0 - (ln.y / 2000.0))
            scores[lot] = scores.get(lot, 0.0) + sc

    if not scores:
        return ""

    # 최고점 lot 반환
    return max(scores.items(), key=lambda x: x[1])[0]


def guess_lot_from_table_text(t: ParsedTable) -> str:
    """
    테이블 텍스트에서 지번 후보를 추정.
    - 권리자/기타사항에 주소(…리 496-10)가 들어가는 케이스가 많아 효과적
    """
    scores: Dict[str, float] = {}
    # 너무 큰 테이블도 있으니, 상단 80행 정도만
    max_r = min(t.n_rows, 80)
    for r in range(max_r):
        row_txt = " ".join((t.grid[r][c] or "") for c in range(t.n_cols))
        for lot in _LOT_SHORT_RE.findall(row_txt):
            sc = 1.0
            if re.search(r"(리|동)\s*" + re.escape(lot), row_txt):
                sc += 2.0
            scores[lot] = scores.get(lot, 0.0) + sc

    if not scores:
        return ""
    return max(scores.items(), key=lambda x: x[1])[0]


def guess_lot_key_for_sec_table(t: ParsedTable, page_lines: List[PageLine]) -> str:
    """
    갑/을구 테이블의 지번 key 추정:
    1) 테이블 텍스트에서 추정
    2) 실패 시 페이지 라인에서 추정
    """
    lot = guess_lot_from_table_text(t)
    if lot:
        return _normalize_lot_key(lot)
    return _normalize_lot_key(guess_lot_from_lines(page_lines))


def guess_lot_key_for_gab_table(t: ParsedTable, page_lines: List[PageLine]) -> str:
    return guess_lot_key_for_sec_table(t, page_lines)


def assign_section_table_to_group(
    groups: List[ParcelGroup],
    t: ParsedTable,
    page_lines: List[PageLine],
) -> Tuple[Optional[ParcelGroup], str, str]:
    """
    갑/을구 테이블을 페이지 범위를 기본 축으로 두고, 지번 추정은 보조 신호로 사용한다.
    반환: (group, assign_mode, guessed_lot)
    """
    guessed_lot = guess_lot_key_for_sec_table(t, page_lines)
    g_by_page = assign_table_to_group(groups, t.page_no)
    g_by_lot = find_group_by_lot_key(groups, guessed_lot) if guessed_lot else None

    if g_by_page is not None and g_by_lot is not None:
        if g_by_page.key == g_by_lot.key:
            return g_by_page, "page_range_lot_agree", guessed_lot
        if g_by_page.key == "UNKNOWN":
            return g_by_lot, "lot_guess_only", guessed_lot
        # 을구/갑구 본문에는 공동담보 등 다른 지번이 자주 섞여서
        # lot guess가 오탐일 수 있으므로, 충돌 시 page range를 우선한다.
        return g_by_page, "page_range_conflict", guessed_lot

    if g_by_page is not None:
        return g_by_page, "page_range", guessed_lot

    if g_by_lot is not None:
        return g_by_lot, "lot_guess_only", guessed_lot

    return None, "unassigned", guessed_lot


# ============================================================
# 8) Excel 생성 (지번별 시트 분리)
# ============================================================
def _safe_sheet_name(name: str, used: set) -> str:
    n = re.sub(r"[\[\]\*:/\\\?]", "_", name).strip()
    if not n:
        n = "sheet"
    n = n[:MAX_SHEETNAME_LEN]
    if n not in used:
        used.add(n)
        return n
    i = 2
    while True:
        cand = f"{n[:MAX_SHEETNAME_LEN-2]}_{i}"[:MAX_SHEETNAME_LEN]
        if cand not in used:
            used.add(cand)
            return cand
        i += 1


def _autosize(ws, max_col: int, max_row: int):
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, max((len(line) for line in s.splitlines()), default=len(s)))
        ws.column_dimensions[letter].width = min(80, max(10, max_len + 2))


def _write_df(ws, df: pd.DataFrame):
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append([r.get(c, "") for c in df.columns])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap
    _autosize(ws, ws.max_column, ws.max_row)


def build_registry_excel_bytes(
    *,
    groups: List[ParcelGroup],
    include_raw_pyo: bool,
    include_raw_gab: bool,
    include_raw_eul: bool,
    merge_cells_on_raw: bool,
) -> bytes:
    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "INDEX"
    ws_index.append(["지번(그룹)", "페이지범위", "표제부 행수", "갑구 행수", "을구 행수", "표제부 테이블 수", "갑구 테이블 수", "을구 테이블 수"])

    used = {"INDEX"}

    # 그룹별 시트
    for g in groups:
        # index row
        ws_index.append([
            g.key,
            f"{g.start_page}-{g.end_page}",
            int(len(g.pyo_df)) if isinstance(g.pyo_df, pd.DataFrame) else 0,
            int(len(g.gab_df)) if isinstance(g.gab_df, pd.DataFrame) else 0,
            int(len(g.eul_df)) if isinstance(g.eul_df, pd.DataFrame) else 0,
            len(g.pyo_tables),
            len(g.gab_tables),
            len(g.eul_tables),
        ])

        # 표제부 시트
        pyo_df = g.pyo_df if isinstance(g.pyo_df, pd.DataFrame) else pd.DataFrame()
        pyo_cols_clean = ["표시번호", "접수", "소재지번", "지목", "면적", "등기원인 및 기타사항"]
        pyo_export = pyo_df[[c for c in pyo_cols_clean if c in pyo_df.columns]].copy() if not pyo_df.empty else pd.DataFrame(columns=pyo_cols_clean)
        ws_pyo = wb.create_sheet(_safe_sheet_name(f"표제부_{g.key}", used))
        if pyo_export.empty:
            ws_pyo.append(["표제부 없음"])
        else:
            _write_df(ws_pyo, pyo_export)

        # 갑구 시트
        gab_df = g.gab_df if isinstance(g.gab_df, pd.DataFrame) else pd.DataFrame()
        gab_cols_clean = ["순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
        gab_export = gab_df[[c for c in gab_cols_clean if c in gab_df.columns]].copy() if not gab_df.empty else pd.DataFrame(columns=gab_cols_clean)
        ws_gab = wb.create_sheet(_safe_sheet_name(f"갑구_{g.key}", used))
        if gab_export.empty:
            ws_gab.append(["갑구 없음"])
        else:
            _write_df(ws_gab, gab_export)

        
        # 을구 시트
        eul_df = g.eul_df if isinstance(g.eul_df, pd.DataFrame) else pd.DataFrame()
        eul_cols_clean = ["순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
        eul_export = eul_df[[c for c in eul_cols_clean if c in eul_df.columns]].copy() if not eul_df.empty else pd.DataFrame(columns=eul_cols_clean)
        ws_eul = wb.create_sheet(_safe_sheet_name(f"을구_{g.key}", used))
        if eul_export.empty:
            ws_eul.append(["을구 없음"])
        else:
            _write_df(ws_eul, eul_export)

# raw tables (선택)
        wrap = Alignment(wrap_text=True, vertical="top")

        def write_raw_table(t: ParsedTable, sheet_prefix: str):
            sname = _safe_sheet_name(f"{sheet_prefix}_{t.table_id}", used)
            ws = wb.create_sheet(sname)
            for r in range(t.n_rows):
                ws.append(t.grid[r])
            for row in ws.iter_rows(min_row=1, max_row=t.n_rows, min_col=1, max_col=t.n_cols):
                for cell in row:
                    cell.alignment = wrap
            if merge_cells_on_raw:
                for (r0, c0, r1, c1) in t.merges:
                    ws.merge_cells(start_row=r0 + 1, start_column=c0 + 1, end_row=r1 + 1, end_column=c1 + 1)
            _autosize(ws, t.n_cols, t.n_rows)

        if include_raw_pyo:
            for t in g.pyo_tables:
                write_raw_table(t, "rawPYO")

        if include_raw_gab:
            for t in g.gab_tables:
                write_raw_table(t, "rawGAB")

        if include_raw_eul:
            for t in g.eul_tables:
                write_raw_table(t, "rawEUL")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# 9) JSON-LD(온톨로지) — 표제부 + 갑구 + 을구
# ============================================================
def build_registry_jsonld(
    *,
    file_name: str,
    file_hash: str,
    groups: List[ParcelGroup],
    base_iri: str = "urn:dovi:",
    generator: str = "DOVI-REGISTRY",
) -> Dict[str, Any]:
    """
    매우 단순한 JSON-LD 그래프:
      Document -> Parcel -> Facts(표제부) + GabEntry(갑구 항목) + EulEntry(을구 항목)
    """
    base = base_iri if base_iri.endswith(("/", "#", ":")) else base_iri + ":"
    doc_id = f"{base}document:{file_hash}"

    context = {
        "@version": 1.1,
        "dovi": "https://example.org/dovi#",
        "schema": "https://schema.org/",
        "xsd": "http://www.w3.org/2001/XMLSchema#",
        "id": "@id",
        "type": "@type",

        "Document": "dovi:Document",
        "Parcel": "dovi:Parcel",
        "Fact": "dovi:Fact",
        "GabEntry": "dovi:GabEntry",
        "EulEntry": "dovi:EulEntry",

        "fileName": "schema:name",
        "fileHash": "dovi:fileHash",
        "createdAt": {"@id": "schema:dateCreated", "@type": "xsd:dateTime"},
        "generator": "schema:generator",

        "mentionsParcel": {"@id": "dovi:mentionsParcel", "@type": "@id"},
        "hasFact": {"@id": "dovi:hasFact", "@type": "@id"},
        "hasGabEntry": {"@id": "dovi:hasGabEntry", "@type": "@id"},
        "hasEulEntry": {"@id": "dovi:hasEulEntry", "@type": "@id"},

        "lot": "dovi:lot",
        "siteAddress": "dovi:siteAddress",

        "field": "dovi:field",
        "value": "dovi:value",

        "rankNo": "dovi:rankNo",
        "purpose": "dovi:purpose",
        "acceptance": "dovi:acceptance",
        "cause": "dovi:cause",
        "holderNote": "dovi:holderNote",
    }

    def now_iso():
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    graph: List[Dict[str, Any]] = []
    doc_node: Dict[str, Any] = {
        "@id": doc_id,
        "@type": "Document",
        "fileName": file_name,
        "fileHash": file_hash,
        "createdAt": now_iso(),
        "generator": generator,
        "mentionsParcel": [],
    }
    graph.append(doc_node)

    def parcel_id(key: str) -> str:
        hid = hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]
        return f"{doc_id}#parcel-{hid}"

    for g in groups:
        pid = parcel_id(g.key)
        doc_node["mentionsParcel"].append(pid)

        # 표제부에서 대표 주소(첫 행) 가져오기
        addr = ""
        if isinstance(g.pyo_df, pd.DataFrame) and not g.pyo_df.empty and "소재지번" in g.pyo_df.columns:
            addr = str(g.pyo_df.iloc[0].get("소재지번", "") or "")

        parcel_node: Dict[str, Any] = {
            "@id": pid,
            "@type": "Parcel",
            "lot": g.key,
            "siteAddress": addr,
            "hasFact": [],
            "hasGabEntry": [],
            "hasEulEntry": [],
        }
        graph.append(parcel_node)

        # 표제부 Facts
        if isinstance(g.pyo_df, pd.DataFrame) and not g.pyo_df.empty:
            for _, row in g.pyo_df.iterrows():
                for field_name in ["표시번호", "접수", "소재지번", "지목", "면적", "등기원인 및 기타사항"]:
                    v = str(row.get(field_name, "") or "").strip()
                    if not v:
                        continue
                    fid = f"{pid}#fact-{hashlib.sha1((field_name+'|'+v).encode('utf-8')).hexdigest()[:10]}"
                    graph.append({"@id": fid, "@type": "Fact", "field": field_name, "value": v})
                    parcel_node["hasFact"].append(fid)

        # 갑구 Entries
        if isinstance(g.gab_df, pd.DataFrame) and not g.gab_df.empty:
            for _, row in g.gab_df.iterrows():
                rank = str(row.get("순위번호", "") or "").strip()
                purpose = str(row.get("등기목적", "") or "").strip()
                acc = str(row.get("접수", "") or "").strip()
                cause = str(row.get("등기원인", "") or "").strip()
                holder = str(row.get("권리자 및 기타사항", "") or "").strip()
                if not (rank or purpose or acc or cause or holder):
                    continue
                eid = f"{pid}#gab-{hashlib.sha1((rank+'|'+purpose+'|'+acc).encode('utf-8')).hexdigest()[:12]}"
                graph.append(
                    {
                        "@id": eid,
                        "@type": "GabEntry",
                        "rankNo": rank,
                        "purpose": purpose,
                        "acceptance": acc,
                        "cause": cause,
                        "holderNote": holder,
                    }
                )
                parcel_node["hasGabEntry"].append(eid)


        # 을구 Entries
        if isinstance(g.eul_df, pd.DataFrame) and not g.eul_df.empty:
            for _, row in g.eul_df.iterrows():
                rank = str(row.get("순위번호", "") or "").strip()
                purpose = str(row.get("등기목적", "") or "").strip()
                acc = str(row.get("접수", "") or "").strip()
                cause = str(row.get("등기원인", "") or "").strip()
                holder = str(row.get("권리자 및 기타사항", "") or "").strip()
                if not (rank or purpose or acc or cause or holder):
                    continue
                eid = f"{pid}#eul-{hashlib.sha1((rank+'|'+purpose+'|'+acc).encode('utf-8')).hexdigest()[:12]}"
                graph.append(
                    {
                        "@id": eid,
                        "@type": "EulEntry",
                        "rankNo": rank,
                        "purpose": purpose,
                        "acceptance": acc,
                        "cause": cause,
                        "holderNote": holder,
                    }
                )
                parcel_node["hasEulEntry"].append(eid)

    return {"@context": context, "@graph": graph}


def make_jsonld_bytes(obj: Dict[str, Any]) -> bytes:
    return json.dumps(obj, ensure_ascii=False, indent=2).encode("utf-8")


# ============================================================
# 10) 전체 처리
# ============================================================
def _parse_rank_token(rank: Any) -> Optional[Tuple[int, int, str]]:
    s = _normalize_rank_text(str(rank or "")) or str(rank or "").strip()
    m = re.match(r"^(\d+)(?:-(\d+))?", s)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2) or 0), s


def _find_rank_jump_warnings(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    순위번호 흐름에서 큰 점프(예: 8-2 → 18)를 디버그용으로 탐지.
    """
    if df is None or df.empty or "순위번호" not in df.columns:
        return []

    rows: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        parsed = _parse_rank_token(row.get("순위번호", ""))
        if parsed is None:
            continue
        main_no, sub_no, rank_raw = parsed
        page_no_raw = row.get("페이지", 0)
        page_no = int(page_no_raw) if str(page_no_raw).strip().isdigit() else 0
        rows.append(
            {
                "main": main_no,
                "sub": sub_no,
                "rank": rank_raw,
                "page": page_no,
                "table_id": str(row.get("table_id", "") or ""),
            }
        )

    if len(rows) < 2:
        return []

    rows.sort(key=lambda x: (x["main"], x["sub"], x["page"], x["table_id"]))
    warnings: List[Dict[str, Any]] = []

    prev = rows[0]
    for cur in rows[1:]:
        is_jump = False
        if cur["main"] > prev["main"] + 1:
            is_jump = True
        elif cur["main"] == prev["main"] and cur["sub"] > prev["sub"] + 1:
            is_jump = True

        if is_jump:
            warnings.append(
                {
                    "from_rank": prev["rank"],
                    "to_rank": cur["rank"],
                    "from_page": prev["page"],
                    "to_page": cur["page"],
                    "from_table_id": prev["table_id"],
                    "to_table_id": cur["table_id"],
                }
            )
        prev = cur

    return warnings


def _rank_main_no(rank: Any) -> Optional[int]:
    s = str(rank or "").strip()
    if not s:
        return None
    m = re.match(r"^(\d+)", s)
    if not m:
        return None
    return int(m.group(1))


def _find_missing_main_ranks(df: pd.DataFrame, *, max_gap: int = 8) -> List[int]:
    if df is None or df.empty or "순위번호" not in df.columns:
        return []

    mains = sorted({m for m in (_rank_main_no(v) for v in df["순위번호"].tolist()) if m is not None})
    if len(mains) < 2:
        return []

    miss: List[int] = []
    for a, b in zip(mains, mains[1:]):
        gap = b - a
        if 1 < gap <= max_gap:
            miss.extend(list(range(a + 1, b)))
    return sorted(set(miss))


def _sort_sec_df_by_rank(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "순위번호" not in df.columns:
        return df

    def rank_key(x: Any) -> Tuple[int, int, str]:
        s = str(x or "").strip()
        m = re.match(r"^(\d+)(?:-(\d+))?", s)
        if not m:
            return (10**9, 0, s)
        return int(m.group(1)), int(m.group(2) or 0), s

    out = df.copy()
    out["_rk"] = out["순위번호"].apply(rank_key)
    out = out.sort_values("_rk").drop(columns=["_rk"], errors="ignore")
    return out


def _recover_rows_from_table_by_missing_mains(
    t: ParsedTable,
    missing_main_set: set,
    *,
    section: str,
    header_row: int = -1,
    col_map: Optional[Dict[str, int]] = None,
) -> pd.DataFrame:
    """
    결번(main rank)만 대상으로 느슨하게 행 복구.
    - 기존 추출 경로에서 놓친 3,4 같은 행을 보완하는 용도
    """
    if not missing_main_set:
        return pd.DataFrame()
    if t.n_rows < 2 or t.n_cols < 2:
        return pd.DataFrame()

    cm = dict(col_map or {})
    c_rank = max(0, min(int(cm.get("rank", 0)), t.n_cols - 1))
    c_purpose = max(0, min(int(cm.get("purpose", min(1, t.n_cols - 1))), t.n_cols - 1))
    c_acc = max(0, min(int(cm.get("acceptance", min(2, t.n_cols - 1))), t.n_cols - 1))
    c_cause = max(0, min(int(cm.get("cause", min(3, t.n_cols - 1))), t.n_cols - 1))
    c_holder = max(0, min(int(cm.get("holder", min(4, t.n_cols - 1))), t.n_cols - 1))

    records: List[Dict[str, Any]] = []
    # 결번 복구는 헤더 오판 가능성을 고려해 테이블 전체 스캔
    start_r = 0
    for r in range(start_r, t.n_rows):
        row = t.grid[r]
        if all((not (row[c] or "").strip()) for c in range(t.n_cols)):
            continue

        rank = _normalize_rank_text(row[c_rank] or "")
        rank_candidates: List[str] = []
        if not rank:
            for c in range(min(3, t.n_cols)):
                rank_candidates.extend(_extract_rank_candidates(row[c] or ""))
            for rk in rank_candidates:
                mn = _rank_main_no(rk)
                if mn in missing_main_set:
                    rank = rk
                    break
            if not rank and rank_candidates:
                rank = rank_candidates[0]
        if not rank:
            for c in range(min(3, t.n_cols)):
                rp, _rest = _split_rank_prefix(row[c] or "")
                if rp:
                    rank = rp
                    break
        if not rank:
            continue
        if _contains_any(_norm(rank), ["순위번호", "갑구", "을구", "표제부"]):
            continue

        main_no = _rank_main_no(rank)
        if main_no is None or main_no not in missing_main_set:
            continue

        row_fields = _extract_sec_row_fields(t, r, header_row=header_row, col_map=cm)
        purpose = row_fields.get("purpose", "")
        acc = row_fields.get("acceptance", "")
        cause = row_fields.get("cause", "")
        holder = row_fields.get("holder", "")

        if not purpose:
            purpose = (row[min(1, t.n_cols - 1)] or "").strip()
        if not acc and t.n_cols >= 3:
            acc = (row[min(2, t.n_cols - 1)] or "").strip()
        if not cause and t.n_cols >= 4:
            cause = (row[min(3, t.n_cols - 1)] or "").strip()

        purpose, acc, cause, holder = _rebalance_sec_fields(
            purpose=purpose,
            acc=acc,
            cause=cause,
            holder=holder,
        )

        ptype = _purpose_type(purpose)
        if section == "을구" and ptype == "gab":
            continue
        if section == "갑구" and ptype == "eul":
            continue

        records.append(
            {
                "페이지": t.page_no,
                "table_id": t.table_id,
                "순위번호": rank,
                "등기목적": purpose,
                "접수": acc,
                "등기원인": cause,
                "권리자 및 기타사항": holder,
            }
        )

    # row 단위에서 놓친 결번을 cell 단위로 재복구 (예: 한 셀에 '3\\n4')
    recovered_mains = {
        m for m in (_rank_main_no(rec.get("순위번호", "")) for rec in records) if m is not None
    }
    unresolved = set(missing_main_set) - recovered_mains
    if unresolved:
        cells_sorted = sorted(t.cells, key=lambda c: (int(c.row), int(c.col)))
        for cell in cells_sorted:
            if not unresolved:
                break
            # 순위번호는 보통 좌측 컬럼에 위치
            if int(cell.col) > max(2, c_rank + 1):
                continue
            for rk in _extract_rank_candidates(cell.text or ""):
                mn = _rank_main_no(rk)
                if mn is None or mn not in unresolved:
                    continue
                rr = int(cell.row)
                if rr < 0 or rr >= t.n_rows:
                    continue
                row = t.grid[rr]
                row_fields = _extract_sec_row_fields(t, rr, header_row=header_row, col_map=cm)
                purpose = row_fields.get("purpose", "")
                acc = row_fields.get("acceptance", "")
                cause = row_fields.get("cause", "")
                holder = row_fields.get("holder", "")
                if not purpose:
                    purpose = (row[min(1, t.n_cols - 1)] or "").strip()
                if not acc and t.n_cols >= 3:
                    acc = (row[min(2, t.n_cols - 1)] or "").strip()
                if not cause and t.n_cols >= 4:
                    cause = (row[min(3, t.n_cols - 1)] or "").strip()

                purpose, acc, cause, holder = _rebalance_sec_fields(
                    purpose=purpose,
                    acc=acc,
                    cause=cause,
                    holder=holder,
                )

                ptype = _purpose_type(purpose)
                if section == "을구" and ptype == "gab":
                    continue
                if section == "갑구" and ptype == "eul":
                    continue

                records.append(
                    {
                        "페이지": t.page_no,
                        "table_id": t.table_id,
                        "순위번호": rk,
                        "등기목적": purpose,
                        "접수": acc,
                        "등기원인": cause,
                        "권리자 및 기타사항": holder,
                    }
                )
                unresolved.discard(mn)
                break

    out = pd.DataFrame(records)
    if out.empty:
        return out
    out = out.drop_duplicates()
    return _sort_sec_df_by_rank(out)


def _recover_rows_from_cells_by_missing_mains(
    t: ParsedTable,
    missing_main_set: set,
    *,
    section: str,
    col_map: Optional[Dict[str, int]] = None,
) -> pd.DataFrame:
    """
    rowIndex가 틀어진 경우를 대비해 cell bbox의 y 위치로 행을 다시 묶어 복구.
    - 셀은 존재하지만 OCR rowIndex가 잘못 잡혀 3,4행이 사라지는 케이스를 노린다.
    """
    if not missing_main_set or not t.cells or t.n_cols < 2:
        return pd.DataFrame()

    cm = dict(col_map or {})
    c_rank = max(0, min(int(cm.get("rank", 0)), t.n_cols - 1))
    c_purpose = max(0, min(int(cm.get("purpose", min(1, t.n_cols - 1))), t.n_cols - 1))
    c_acc = max(0, min(int(cm.get("acceptance", min(2, t.n_cols - 1))), t.n_cols - 1))
    c_cause = max(0, min(int(cm.get("cause", min(3, t.n_cols - 1))), t.n_cols - 1))
    c_holder = max(0, min(int(cm.get("holder", min(4, t.n_cols - 1))), t.n_cols - 1))

    cells = [pc for pc in t.cells if pc.bbox and (pc.text or "").strip()]
    if not cells:
        return pd.DataFrame()

    heights = sorted((float(pc.bbox[3]) - float(pc.bbox[1])) for pc in cells if pc.bbox)
    if heights:
        h_med = heights[len(heights) // 2]
        y_thresh = max(12.0, min(28.0, h_med * 0.7))
    else:
        y_thresh = 18.0

    line_items: List[Tuple[float, float, ParsedCell]] = []
    for pc in cells:
        x0, y0, x1, y1 = pc.bbox or (0.0, 0.0, 0.0, 0.0)
        y_mid = (float(y0) + float(y1)) / 2.0
        line_items.append((y_mid, float(x0), pc))

    line_items.sort(key=lambda x: (x[0], x[1]))

    clusters: List[List[Tuple[float, float, ParsedCell]]] = []
    cur: List[Tuple[float, float, ParsedCell]] = []
    cur_y: Optional[float] = None

    def flush_cluster():
        nonlocal cur, cur_y
        if cur:
            clusters.append(cur)
        cur = []
        cur_y = None

    for y_mid, x0, pc in line_items:
        if cur and cur_y is not None and abs(y_mid - cur_y) > y_thresh:
            flush_cluster()
        cur.append((y_mid, x0, pc))
        if cur_y is None:
            cur_y = y_mid
        else:
            cur_y = (cur_y * (len(cur) - 1) + y_mid) / len(cur)
    flush_cluster()

    records: List[Dict[str, Any]] = []
    seen_ranks: set = set()

    for cluster in clusters:
        row = ["" for _ in range(t.n_cols)]
        by_col: Dict[int, List[Tuple[float, str]]] = {}

        for _y_mid, x0, pc in cluster:
            if not (0 <= int(pc.col) < t.n_cols):
                continue
            txt = (pc.text or "").strip()
            if not txt:
                continue
            by_col.setdefault(int(pc.col), []).append((x0, txt))

        for c, parts in by_col.items():
            parts.sort(key=lambda x: x[0])
            uniq_parts: List[str] = []
            seen_parts: set = set()
            for _x0, txt in parts:
                if txt in seen_parts:
                    continue
                seen_parts.add(txt)
                uniq_parts.append(txt)
            row[c] = "\n".join(uniq_parts).strip()

        if not any((v or "").strip() for v in row):
            continue

        rank = _normalize_rank_text(row[c_rank] or "")
        if not rank:
            probe_cols = sorted({c_rank, 0, 1, 2} & set(range(t.n_cols)))
            for c in probe_cols:
                probe = _normalize_rank_text(row[c] or "")
                if probe:
                    rank = probe
                    break
                for rk in _extract_rank_candidates(row[c] or ""):
                    mn = _rank_main_no(rk)
                    if mn in missing_main_set:
                        rank = rk
                        break
                if rank:
                    break
                rk2, _rest = _split_rank_prefix(row[c] or "")
                if rk2:
                    rank = rk2
                    break

        if not rank or rank in seen_ranks:
            continue
        if _contains_any(_norm(rank), ["순위번호", "갑구", "을구", "표제부"]):
            continue

        main_no = _rank_main_no(rank)
        if main_no is None or main_no not in missing_main_set:
            continue

        purpose = join_cols(row, c_purpose, c_acc, sep="\n") if c_purpose < c_acc else (row[c_purpose] or "").strip()
        acc = join_cols(row, c_acc, c_cause, sep="\n") if c_acc < c_cause else (row[c_acc] or "").strip()
        cause = join_cols(row, c_cause, c_holder, sep="\n") if c_cause < c_holder else (row[c_cause] or "").strip()
        holder = join_cols(row, c_holder, t.n_cols, sep="\n")

        if not purpose and t.n_cols >= 2:
            purpose = (row[min(1, t.n_cols - 1)] or "").strip()
        if not acc and t.n_cols >= 3:
            acc = (row[min(2, t.n_cols - 1)] or "").strip()
        if not cause and t.n_cols >= 4:
            cause = (row[min(3, t.n_cols - 1)] or "").strip()

        purpose, acc, cause, holder = _rebalance_sec_fields(
            purpose=purpose,
            acc=acc,
            cause=cause,
            holder=holder,
        )

        ptype = _purpose_type(" ".join([purpose, holder]).strip())
        if section == "을구" and ptype == "gab":
            continue
        if section == "갑구" and ptype == "eul":
            continue

        records.append(
            {
                "페이지": t.page_no,
                "table_id": f"{t.table_id}_cellcluster",
                "순위번호": rank,
                "등기목적": purpose,
                "접수": acc,
                "등기원인": cause,
                "권리자 및 기타사항": holder,
            }
        )
        seen_ranks.add(rank)

    out = pd.DataFrame(records)
    if out.empty:
        return out
    out = out.drop_duplicates()
    return _sort_sec_df_by_rank(out)


def _extract_date_phrases(text: str) -> List[str]:
    if not text:
        return []
    # 예: 2005년5월20일 제65032호 / 2005년5월20일 해지
    pat = re.compile(
        rf"{SEC_DATE_RE_TEXT}(?:\s*(?:제\s*)?\d+\s*호)?(?:\s*{SEC_CAUSE_ACTION_RE_TEXT})?"
    )
    out = [re.sub(r"\s+", " ", m.group(0)).strip() for m in pat.finditer(text)]
    # 순서 유지 unique
    seen = set()
    uniq: List[str] = []
    for x in out:
        if x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    return uniq


def _purpose_from_rank_line(rank: str, line_text: str) -> str:
    s = (line_text or "").strip()
    if not s:
        return ""
    # rank prefix 제거
    s = re.sub(r"^\s*" + re.escape(rank) + r"\s*", "", s).strip()
    if not s:
        return ""
    # 날짜 구문 전까지를 등기목적으로 간주
    m = re.search(r"\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일", s)
    if m:
        s = s[:m.start()].strip()
    return s


def _leading_rank_from_line(line_text: str) -> str:
    """
    line 맨 앞의 순위번호를 느슨하게 추출.
    - '3', '3-1', '3(전3)', '3번', '3.' 형태 허용
    - 연도/금액 오탐(예: 2005) 방지를 위해 main <= 300 제한
    """
    s = (line_text or "").strip()
    if not s:
        return ""

    m = re.match(r"^\s*(\d{1,4}(?:-\d{1,4})?(?:\s*\([^)]*\))?)", s)
    if m:
        rk = _normalize_rank_text(m.group(1) or "")
        mn = _rank_main_no(rk)
        if rk and mn is not None and mn <= 300:
            return rk

    cands = _extract_rank_candidates(s[:16])
    for rk in cands:
        mn = _rank_main_no(rk)
        if mn is not None and mn <= 300:
            return rk
    return ""


def _recover_rows_from_page_lines_by_missing_mains(
    all_page_lines: Dict[int, List[PageLine]],
    pages: List[int],
    missing_main_set: set,
    *,
    section: str,
) -> pd.DataFrame:
    """
    fields 기반(line) 결번 복구.
    - table 복구 이후에도 남는 결번(예: 3,4)을 page_lines에서 다시 탐색.
    """
    if not missing_main_set:
        return pd.DataFrame()

    records: List[Dict[str, Any]] = []

    for p in pages:
        lines = all_page_lines.get(p, [])
        if not lines:
            continue

        txts = [str(ln.text or "").strip() for ln in lines]
        n = len(txts)
        i = 0
        while i < n:
            line = txts[i]
            rank = _leading_rank_from_line(line)
            if not rank:
                i += 1
                continue

            main_no = _rank_main_no(rank)
            if not rank or main_no is None or main_no not in missing_main_set:
                i += 1
                continue

            # 현재 rank 블록 수집 (다음 rank 시작 전까지)
            j = i + 1
            while j < n:
                if _leading_rank_from_line(txts[j]):
                    break
                j += 1

            block_lines = [t for t in txts[i:j] if t]
            block_text = "\n".join(block_lines).strip()
            if not block_text:
                i = j
                continue

            purpose = _purpose_from_rank_line(rank, block_lines[0] if block_lines else "")
            date_phrases = _extract_date_phrases(block_text)
            acc = date_phrases[0] if len(date_phrases) >= 1 else ""
            cause = date_phrases[1] if len(date_phrases) >= 2 else ""

            holder = block_text
            holder = re.sub(r"^\s*" + re.escape(rank) + r"\s*", "", holder).strip()
            if purpose:
                holder = holder.replace(purpose, "", 1).strip()
            if acc:
                holder = holder.replace(acc, "", 1).strip()
            if cause:
                holder = holder.replace(cause, "", 1).strip()
            holder = re.sub(r"\n{2,}", "\n", holder).strip()

            purpose, acc, cause, holder = _rebalance_sec_fields(
                purpose=purpose,
                acc=acc,
                cause=cause,
                holder=holder,
            )

            ptype = _purpose_type(purpose + " " + holder)
            if section == "을구" and ptype == "gab":
                i = j
                continue
            if section == "갑구" and ptype == "eul":
                i = j
                continue

            records.append(
                {
                    "페이지": p,
                    "table_id": f"fields_p{p}_r{i+1}",
                    "순위번호": rank,
                    "등기목적": purpose,
                    "접수": acc,
                    "등기원인": cause,
                    "권리자 및 기타사항": holder,
                }
            )

            i = j

    out = pd.DataFrame(records)
    if out.empty:
        return out
    out = out.drop_duplicates()
    return _sort_sec_df_by_rank(out)


def _extract_main_nos_from_rankish_text(text: str) -> List[int]:
    """
    rank처럼 보이는 텍스트에서 main rank 번호만 추출.
    - 디버그 provenance용이라 엄격한 파싱보다 '있었는지' 판별에 초점을 둔다.
    """
    s = (text or "").strip()
    if not s:
        return []

    mains: List[int] = []

    norm = _normalize_rank_text(s)
    if norm:
        mn = _rank_main_no(norm)
        if mn is not None and mn <= 300:
            mains.append(mn)

    rp, _rest = _split_rank_prefix(s)
    if rp:
        mn = _rank_main_no(rp)
        if mn is not None and mn <= 300:
            mains.append(mn)

    for rk in _extract_rank_candidates(s):
        mn = _rank_main_no(rk)
        if mn is not None and mn <= 300:
            mains.append(mn)

    return sorted(set(mains))


def _is_potential_eul_source_table(
    t: ParsedTable,
    section_hint_by_table: Dict[str, str],
    page_lines: List[PageLine],
) -> bool:
    """
    을구 provenance 진단 시 참고할 만한 테이블만 선별.
    - 갑구로 판단된 테이블은 제외
    - 을구/연속표 후보는 포함
    """
    sec_hint = section_hint_by_table.get(t.table_id, "unknown")
    if sec_hint == "갑구":
        return False
    if sec_hint == "을구":
        return True

    by_label = classify_gab_or_eul(t, page_lines)
    if by_label == "갑구":
        return False
    if by_label == "을구":
        return True

    return _looks_like_sec_continuation_table(t)


def _collect_missing_rank_source_provenance(
    *,
    group_key: str,
    target_mains: List[int],
    final_eul_df: pd.DataFrame,
    candidate_tables: List[ParsedTable],
    pages_in_group: List[int],
    all_page_lines: Dict[int, List[PageLine]],
) -> List[Dict[str, Any]]:
    """
    결번 순위번호가 OCR 원응답의 어느 층위에 있었는지 추적.
    - grid: 구조화된 table row/col
    - cells: raw table cell 텍스트
    - fields: 일반 OCR line
    """
    if not target_mains:
        return []

    final_mains = {
        m for m in (_rank_main_no(v) for v in final_eul_df.get("순위번호", pd.Series(dtype=str)).tolist()) if m is not None
    }

    out: List[Dict[str, Any]] = []
    table_col_limit = 2  # rank가 놓일 가능성이 높은 좌측 컬럼만 본다.

    for target in sorted(set(target_mains)):
        grid_tables: List[str] = []
        grid_pages: set = set()
        cell_tables: List[str] = []
        cell_pages: set = set()
        field_pages: set = set()

        for t in candidate_tables:
            grid_hit_here = False
            for r in range(t.n_rows):
                for c in range(min(table_col_limit, t.n_cols)):
                    if target in _extract_main_nos_from_rankish_text(t.grid[r][c] or ""):
                        grid_tables.append(t.table_id)
                        grid_pages.add(int(t.page_no))
                        grid_hit_here = True
                        break
                if grid_hit_here:
                    break

            for cell in t.cells:
                if int(cell.col) >= table_col_limit:
                    continue
                if target in _extract_main_nos_from_rankish_text(cell.text or ""):
                    cell_tables.append(t.table_id)
                    cell_pages.add(int(t.page_no))
                    break

        for p in pages_in_group:
            lines = all_page_lines.get(p, [])
            if not lines:
                continue
            for ln in lines:
                rk = _leading_rank_from_line(ln.text or "")
                if _rank_main_no(rk) == target:
                    field_pages.add(int(p))
                    break

        grid_hit = len(grid_tables) > 0
        cell_hit = len(cell_tables) > 0
        fields_hit = len(field_pages) > 0
        final_hit = target in final_mains

        if final_hit:
            diagnosis = "최종 결과 반영됨"
        elif not cell_hit and not fields_hit:
            diagnosis = "CLOVA 원응답 누락 의심"
        elif (cell_hit or fields_hit) and not grid_hit:
            diagnosis = "raw OCR 존재, 구조화 단계 누락 의심"
        else:
            diagnosis = "구조화 OCR 존재, 최종 파싱 누락 의심"

        out.append(
            {
                "group_key": group_key,
                "rank_main": target,
                "present_in_final_eul_df": final_hit,
                "grid_hit": grid_hit,
                "grid_hit_pages": sorted(grid_pages),
                "grid_hit_tables": sorted(set(grid_tables)),
                "cell_hit": cell_hit,
                "cell_hit_pages": sorted(cell_pages),
                "cell_hit_tables": sorted(set(cell_tables)),
                "fields_hit": fields_hit,
                "fields_hit_pages": sorted(field_pages),
                "diagnosis": diagnosis,
            }
        )

    return out


def process_pdf(
    file_bytes: bytes,
    api_url: str,
    secret_key: str,
    *,
    pages_per_request: int,
    lang: str,
    progress_cb: Optional[Callable[[int, int, int, int], None]] = None,
    debug: bool = False,
) -> Tuple[List[ParcelGroup], int, Dict[str, Any]]:
    """
    반환:
      - groups: 지번별 ParcelGroup 리스트(표제부+갑구+을구)
      - total_pages
      - debug_info: 페이지/테이블 추적 메타
    """
    pages_per_request = max(1, min(int(pages_per_request), MAX_PAGES_PER_REQUEST))
    total_pages = get_pdf_total_pages(file_bytes)
    chunks = split_pdf_into_chunks(file_bytes, pages_per_request)

    debug_info: Dict[str, Any] = {
        "total_pages": total_pages,
        "ocr_chunks": [],
        "land_registry_pages": [],
        "land_page_trace": [],
        "table_counts": {},
        "pyo_candidates": [],
        "pyo_empty_tables": [],
        "pyo_accepted_tables": [],
        "section_candidates": [],
        "section_rejected_tables": [],
        "section_skipped_tables": [],
        "section_accepted_tables": [],
        "eul_gap_recovery": [],
        "eul_fields_recovery": [],
        "eul_source_provenance": [],
        "group_ranges": [],
        "rank_jump_warnings": [],
    }

    all_tables: List[ParsedTable] = []
    all_page_lines: Dict[int, List[PageLine]] = {}

    for i, (chunk_pdf, start_p, end_p) in enumerate(chunks, start=1):
        if progress_cb:
            progress_cb(i, len(chunks), start_p, end_p)

        res = call_naver_ocr_table(chunk_pdf, api_url, secret_key, lang=lang)
        if not res.get("ok"):
            raise RuntimeError(
                f"OCR 실패 (pages {start_p}-{end_p})\n"
                f"status={res.get('status_code')}\n{res.get('text') or res.get('error')}"
            )
        ocr_json = res.get("json")
        if not ocr_json:
            raise RuntimeError(f"OCR JSON 파싱 실패 (pages {start_p}-{end_p})\n{res.get('text')}")

        page_numbers = list(range(start_p, end_p + 1))
        tables, page_lines = parse_tables_and_lines_from_ocr_json(ocr_json, page_numbers=page_numbers)
        all_tables.extend(tables)
        all_page_lines.update(page_lines)
        if debug:
            debug_info["ocr_chunks"].append(
                {
                    "chunk_index": i,
                    "start_page": start_p,
                    "end_page": end_p,
                    "table_count": len(tables),
                }
            )

    if progress_cb:
        progress_cb(len(chunks), len(chunks), 0, 0)

    # --------------------
    # ✅ 토지 등기 관련 페이지만 선별 (토지이용계획확인서/토지대장/건물 등기 등은 제외)
    # --------------------
    land_page_trace: Optional[List[Dict[str, Any]]] = [] if debug else None
    land_registry_pages = compute_land_registry_pages(all_page_lines, total_pages, trace=land_page_trace)
    if debug:
        debug_info["land_registry_pages"] = land_registry_pages
        debug_info["land_page_trace"] = land_page_trace if land_page_trace is not None else []

    if land_registry_pages:
        land_page_set = set(land_registry_pages)
        tables_for_registry = [t for t in all_tables if t.page_no in land_page_set]
    else:
        # 토지 등기 페이지 판별 실패 시(극단적 OCR 실패)에는 전체를 사용(최후의 fallback)
        tables_for_registry = all_tables

    if debug:
        debug_info["table_counts"] = {
            "all_tables": len(all_tables),
            "registry_tables": len(tables_for_registry),
        }

    # 그룹 페이지 범위는 '토지 등기' 마지막 페이지 기준이 더 자연스러움
    last_registry_page = max(land_registry_pages) if land_registry_pages else total_pages

    # --------------------
    # 표제부 추출
    # --------------------
    pyo_candidates = find_pyo_tables(tables_for_registry)
    pyo_items: List[Tuple[ParsedTable, pd.DataFrame]] = []
    for (t, header_row, col_map) in pyo_candidates:
        if debug:
            debug_info["pyo_candidates"].append(
                {"page": t.page_no, "table_id": t.table_id, "header_row": header_row}
            )
        df = extract_pyo_records_from_table(t, header_row, col_map)
        if df.empty:
            if debug:
                debug_info["pyo_empty_tables"].append({"page": t.page_no, "table_id": t.table_id})
            continue
        pyo_items.append((t, df))
        if debug:
            debug_info["pyo_accepted_tables"].append(
                {"page": t.page_no, "table_id": t.table_id, "row_count": int(len(df))}
            )

    groups = group_parcels_from_pyo(pyo_items, total_pages=last_registry_page)

    # 표제부가 하나도 없으면 UNKNOWN 그룹 하나 생성
    if not groups:
        groups = [ParcelGroup(key="UNKNOWN", start_page=1, end_page=last_registry_page)]

    # 그룹에 표제부 테이블/DF 할당
    # (group_parcels_from_pyo가 이미 할당했지만, UNKNOWN 케이스 대비)
    if groups and pyo_items:
        by_key = {g.key: g for g in groups}
        for t, df in pyo_items:
            key = _pick_group_key_from_pyo_df(df, fallback=t.table_id)
            g = by_key.get(key)
            if g is None:
                # 새 그룹 생성(예외)
                g = ParcelGroup(key=key, start_page=t.page_no, end_page=last_registry_page)
                groups.append(g)
                by_key[key] = g
            g.pyo_tables.append(t)
            g.pyo_df = pd.concat([g.pyo_df, df], ignore_index=True) if isinstance(g.pyo_df, pd.DataFrame) and not g.pyo_df.empty else df

    lot_guess_by_table: Dict[str, str] = {
        t.table_id: guess_lot_key_for_sec_table(t, all_page_lines.get(t.page_no, []))
        for t in tables_for_registry
    }

    # --------------------
    # 갑/을구 추출
    # --------------------
    sec_candidates = find_gab_tables(tables_for_registry)
    if debug:
        debug_info["section_candidates"] = [
            {"page": t.page_no, "table_id": t.table_id, "header_row": header_row}
            for (t, header_row, _col_map) in sec_candidates
        ]
        sec_candidate_ids = {t.table_id for (t, _header_row, _col_map) in sec_candidates}
        pyo_table_ids = {t.table_id for (t, _df) in pyo_items}
        rejected_rows: List[Dict[str, Any]] = []
        for t in tables_for_registry:
            if t.table_id in sec_candidate_ids:
                continue
            meta = _analyze_sec_table_candidate(t)
            reason = "weak_sec_signal"
            if t.table_id in pyo_table_ids or bool(meta.get("pyo_like")):
                reason = "pyo_like_table"
            elif int(meta.get("n_rows", 0)) < 2 or int(meta.get("n_cols", 0)) < 3:
                reason = "table_too_small"
            rejected_rows.append(
                {
                    "page": t.page_no,
                    "table_id": t.table_id,
                    "n_rows": int(meta.get("n_rows", 0)),
                    "n_cols": int(meta.get("n_cols", 0)),
                    "header_row": int(meta.get("header_row", -1)),
                    "header_hits": ",".join(meta.get("header_hits", []) or []),
                    "nonempty_rows": int(meta.get("nonempty_rows", 0)),
                    "rank_like_rows": int(meta.get("rank_like_rows", 0)),
                    "purpose_keyword_rows": int(meta.get("purpose_keyword_rows", 0)),
                    "date_like_rows": int(meta.get("date_like_rows", 0)),
                    "holder_like_rows": int(meta.get("holder_like_rows", 0)),
                    "continuation_score": int(meta.get("continuation_score", 0)),
                    "looks_like_continuation": bool(meta.get("looks_like_continuation", False)),
                    "reason": reason,
                }
            )
        debug_info["section_rejected_tables"] = rejected_rows

    # 그룹별 DF 누적
    gab_df_by_group: Dict[str, List[pd.DataFrame]] = {g.key: [] for g in groups}
    eul_df_by_group: Dict[str, List[pd.DataFrame]] = {g.key: [] for g in groups}

    # 처리 순서 안정화: (page_no, bbox_y0, table_id)
    def _tbl_sort_key(x: Tuple[ParsedTable, int, Dict[str, int]]) -> Tuple[int, float, str]:
        t, header_row, col_map = x
        y0 = 0.0
        if t.bbox:
            y0 = float(t.bbox[1])
        else:
            ys = [float(c.bbox[1]) for c in t.cells if c.bbox]
            if ys:
                y0 = min(ys)
        return (int(t.page_no), y0, str(t.table_id))

    sec_candidates.sort(key=_tbl_sort_key)
    sec_colmap_by_id: Dict[str, Dict[str, Any]] = {}
    section_hint_by_table: Dict[str, str] = {}

    for (t, header_row, col_map) in sec_candidates:
        sec_colmap_by_id[t.table_id] = {"header_row": header_row, "col_map": dict(col_map)}
        # 1차: 페이지 라벨(갑구/을구) 기반 분류
        section_by_label = classify_gab_or_eul(t, all_page_lines.get(t.page_no, []))
        section_by_kw = guess_section_by_purpose_keywords(t, header_row, col_map)
        section = section_by_label

        # 2차: 라벨이 안 잡히면 등기목적 키워드로 추정
        if section == "unknown":
            section = section_by_kw

        # 3차: 라벨이 잡힌 경우에는 라벨을 우선한다.
        # (을구에서 '근저당권이전/변경' 같은 키워드가 갑구로 오판되는 것을 방지)
        if section_by_label in ("갑구", "을구"):
            section = section_by_label
        section_hint_by_table[t.table_id] = section

        if section not in ("갑구", "을구"):
            if debug:
                debug_info["section_skipped_tables"].append(
                    {
                        "page": t.page_no,
                        "table_id": t.table_id,
                        "section_by_label": section_by_label,
                        "section_by_kw": section_by_kw,
                        "reason": "unknown_section",
                    }
                )
            continue

        if section == "갑구":
            df = extract_gab_records_from_table(t, header_row, col_map)
        else:
            df = extract_eul_records_from_table(t, header_row, col_map)

        if df.empty:
            if debug:
                debug_info["section_skipped_tables"].append(
                    {
                        "page": t.page_no,
                        "table_id": t.table_id,
                        "section_by_label": section_by_label,
                        "section_by_kw": section_by_kw,
                        "final_section": section,
                        "reason": "empty_records",
                    }
                )
            continue

        # ✅ 우선: 지번 추정 기반 할당, 실패 시 페이지 범위 fallback
        g, group_assign_mode, guessed_lot = assign_section_table_to_group(
            groups,
            t,
            all_page_lines.get(t.page_no, []),
        )

        if g is None:
            # 마지막 fallback: UNKNOWN
            g = next((x for x in groups if x.key == "UNKNOWN"), None)
            if g is None:
                g = ParcelGroup(key="UNKNOWN", start_page=1, end_page=last_registry_page)
                groups.append(g)
                gab_df_by_group.setdefault(g.key, [])
                eul_df_by_group.setdefault(g.key, [])
            group_assign_mode = "unknown_fallback"

        if section == "갑구":
            g.gab_tables.append(t)
            gab_df_by_group.setdefault(g.key, []).append(df)
        else:
            g.eul_tables.append(t)
            eul_df_by_group.setdefault(g.key, []).append(df)

        if debug:
            debug_info["section_accepted_tables"].append(
                {
                    "page": t.page_no,
                    "table_id": t.table_id,
                        "section_by_label": section_by_label,
                        "section_by_kw": section_by_kw,
                        "final_section": section,
                        "group_key": g.key,
                        "lot_guess": guessed_lot,
                        "group_assign_mode": group_assign_mode,
                        "row_count": int(len(df)),
                    }
                )

    # 그룹별 DF finalize
    for g in groups:
        # 갑구
        gdfs = gab_df_by_group.get(g.key) or []
        if gdfs:
            gab_df = pd.concat(gdfs, ignore_index=True).drop_duplicates()
            cols = ["페이지", "table_id", "순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
            g.gab_df = gab_df[[c for c in cols if c in gab_df.columns]]
        else:
            g.gab_df = pd.DataFrame(columns=["순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"])

        # 을구
        edfs = eul_df_by_group.get(g.key) or []
        if edfs:
            eul_df = pd.concat(edfs, ignore_index=True).drop_duplicates()
            cols = ["페이지", "table_id", "순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
            g.eul_df = eul_df[[c for c in cols if c in eul_df.columns]]
        else:
            g.eul_df = pd.DataFrame(columns=["순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"])

    # --------------------
    # 을구 결번 복구(post-recovery)
    # --------------------
    for g in groups:
        if not isinstance(g.eul_df, pd.DataFrame) or g.eul_df.empty:
            continue

        missing_mains = _find_missing_main_ranks(g.eul_df)
        if not missing_mains:
            continue

        missing_set = set(missing_mains)
        recovered_parts: List[pd.DataFrame] = []
        cell_cluster_recovered_rows = 0
        cell_cluster_recovered_ranks: List[str] = []
        fields_recovered_rows = 0
        fields_recovered_ranks: List[str] = []
        pages_in_group = sorted({p for p in land_registry_pages if g.start_page <= p <= g.end_page})
        if not pages_in_group:
            pages_in_group = list(range(g.start_page, g.end_page + 1))

        group_lot_key = _normalize_lot_key(g.key)
        group_has_lot_key = _is_probable_lot_key(group_lot_key)
        cand_tables = []
        for t in tables_for_registry:
            guessed_lot = lot_guess_by_table.get(t.table_id, "")
            page_match = g.start_page <= t.page_no <= g.end_page
            lot_match = group_has_lot_key and guessed_lot == group_lot_key
            # page range 안의 테이블은 항상 보되, 지번 추정이 맞는 테이블은 범위 밖이어도 보조 후보로 포함
            if page_match or lot_match:
                cand_tables.append(t)
        cand_tables.sort(key=lambda t: (int(t.page_no), float(t.bbox[1]) if t.bbox else 0.0, str(t.table_id)))
        pages_in_group = sorted(set(pages_in_group) | {int(t.page_no) for t in cand_tables})

        for t in cand_tables:
            if not missing_set:
                break

            sec_hint = section_hint_by_table.get(t.table_id, "unknown")

            # 힌트가 불명확하면 라벨/연속표 형태를 이용해 을구 복구 대상으로만 제한
            if sec_hint == "unknown":
                by_label = classify_gab_or_eul(t, all_page_lines.get(t.page_no, []))
                if by_label == "갑구":
                    continue
                if by_label == "을구":
                    sec_hint = "을구"
                else:
                    # OCR이 3~4행만 별도 테이블로 쪼개는 경우를 위해,
                    # unknown이라도 결번 복구 시도는 허용한다.
                    sec_hint = "을구"

            meta = sec_colmap_by_id.get(t.table_id, {})
            header_row = int(meta.get("header_row", -1))
            cmap = meta.get("col_map", {}) if isinstance(meta.get("col_map", {}), dict) else {}

            rec = _recover_rows_from_table_by_missing_mains(
                t,
                missing_set,
                section="을구",
                header_row=header_row,
                col_map=cmap,
            )
            if rec.empty:
                continue

            recovered_parts.append(rec)
            recovered_main = {
                m for m in (_rank_main_no(v) for v in rec.get("순위번호", pd.Series(dtype=str)).tolist()) if m is not None
            }
            missing_set -= recovered_main

            if missing_set:
                rec_cells = _recover_rows_from_cells_by_missing_mains(
                    t,
                    missing_set,
                    section="을구",
                    col_map=cmap,
                )
                if not rec_cells.empty:
                    recovered_parts.append(rec_cells)
                    cell_cluster_recovered_rows += int(len(rec_cells))
                    cell_cluster_recovered_ranks.extend(
                        [str(x).strip() for x in rec_cells.get("순위번호", pd.Series(dtype=str)).tolist() if str(x).strip()]
                    )
                    recovered_main_cells = {
                        m for m in (_rank_main_no(v) for v in rec_cells.get("순위번호", pd.Series(dtype=str)).tolist()) if m is not None
                    }
                    missing_set -= recovered_main_cells

        # table row 복구가 비어도, cell bbox만으로는 잡히는 케이스가 있어 한 번 더 시도
        if missing_set:
            for t in cand_tables:
                if not missing_set:
                    break
                sec_hint = section_hint_by_table.get(t.table_id, "unknown")
                if sec_hint == "갑구":
                    continue
                if sec_hint == "unknown":
                    by_label = classify_gab_or_eul(t, all_page_lines.get(t.page_no, []))
                    if by_label == "갑구":
                        continue
                meta = sec_colmap_by_id.get(t.table_id, {})
                cmap = meta.get("col_map", {}) if isinstance(meta.get("col_map", {}), dict) else {}
                rec_cells = _recover_rows_from_cells_by_missing_mains(
                    t,
                    missing_set,
                    section="을구",
                    col_map=cmap,
                )
                if rec_cells.empty:
                    continue
                recovered_parts.append(rec_cells)
                cell_cluster_recovered_rows += int(len(rec_cells))
                cell_cluster_recovered_ranks.extend(
                    [str(x).strip() for x in rec_cells.get("순위번호", pd.Series(dtype=str)).tolist() if str(x).strip()]
                )
                recovered_main_cells = {
                    m for m in (_rank_main_no(v) for v in rec_cells.get("순위번호", pd.Series(dtype=str)).tolist()) if m is not None
                }
                missing_set -= recovered_main_cells

        # table/cell 복구로도 남는 결번은 fields(page_lines) 기반으로 추가 복구
        if missing_set:
            rec_fields = _recover_rows_from_page_lines_by_missing_mains(
                all_page_lines=all_page_lines,
                pages=pages_in_group,
                missing_main_set=set(missing_set),
                section="을구",
            )
            if not rec_fields.empty:
                recovered_parts.append(rec_fields)
                fields_recovered_rows = int(len(rec_fields))
                fields_recovered_ranks = [
                    str(x).strip()
                    for x in rec_fields.get("순위번호", pd.Series(dtype=str)).tolist()
                    if str(x).strip()
                ]
                rec_main_fields = {
                    m for m in (_rank_main_no(v) for v in rec_fields.get("순위번호", pd.Series(dtype=str)).tolist()) if m is not None
                }
                missing_set -= rec_main_fields

        if recovered_parts:
            base = g.eul_df if isinstance(g.eul_df, pd.DataFrame) else pd.DataFrame()
            merged = pd.concat([base] + recovered_parts, ignore_index=True).drop_duplicates()
            merged = _sort_sec_df_by_rank(merged)
            cols = ["페이지", "table_id", "순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
            g.eul_df = merged[[c for c in cols if c in merged.columns]]

        if debug:
            provenance_tables = [
                t for t in cand_tables
                if _is_potential_eul_source_table(t, section_hint_by_table, all_page_lines.get(t.page_no, []))
            ]
            debug_info["eul_source_provenance"].extend(
                _collect_missing_rank_source_provenance(
                    group_key=g.key,
                    target_mains=missing_mains,
                    final_eul_df=g.eul_df if isinstance(g.eul_df, pd.DataFrame) else pd.DataFrame(),
                    candidate_tables=provenance_tables,
                    pages_in_group=pages_in_group,
                    all_page_lines=all_page_lines,
                )
            )
            debug_info["eul_gap_recovery"].append(
                {
                    "group_key": g.key,
                    "before_missing_main_ranks": missing_mains,
                    "remaining_missing_main_ranks": sorted(missing_set),
                    "recovered_parts": len(recovered_parts),
                    "candidate_tables_in_range": len(cand_tables),
                    "cell_cluster_recovered_rows": cell_cluster_recovered_rows,
                    "cell_cluster_recovered_ranks": sorted(set(cell_cluster_recovered_ranks)),
                }
            )
            debug_info["eul_fields_recovery"].append(
                {
                    "group_key": g.key,
                    "fields_recovered_rows": fields_recovered_rows,
                    "fields_recovered_ranks": fields_recovered_ranks,
                    "remaining_missing_main_ranks_after_fields": sorted(missing_set),
                }
            )

    # 표제부 DF 중복 제거
    for g in groups:
        if isinstance(g.pyo_df, pd.DataFrame) and not g.pyo_df.empty:
            g.pyo_df = g.pyo_df.drop_duplicates()

    # --------------------
    # 그룹 범위 재계산(표제부를 못 찾은 지번이 갑구에서 추가된 경우 대비)
    # --------------------
    # 지번 추정 기반 할당을 쓰면 그룹 페이지 범위가 겹칠 수 있으므로,
    # 다른 그룹의 start_page로 끊지 않고 각 그룹의 실제 min/max page를 사용한다.
    for g in groups:
        pages = []
        pages.extend([t.page_no for t in g.pyo_tables])
        pages.extend([t.page_no for t in g.gab_tables])
        pages.extend([t.page_no for t in g.eul_tables])
        if pages:
            g.start_page = min(pages)
            g.end_page = max(pages)

    groups.sort(key=lambda x: x.start_page)

    if debug:
        debug_info["group_ranges"] = [
            {
                "group_key": g.key,
                "start_page": g.start_page,
                "end_page": g.end_page,
                "pyo_rows": int(len(g.pyo_df)) if isinstance(g.pyo_df, pd.DataFrame) else 0,
                "gab_rows": int(len(g.gab_df)) if isinstance(g.gab_df, pd.DataFrame) else 0,
                "eul_rows": int(len(g.eul_df)) if isinstance(g.eul_df, pd.DataFrame) else 0,
            }
            for g in groups
        ]

        rank_jump_warnings: List[Dict[str, Any]] = []
        for g in groups:
            gab_w = _find_rank_jump_warnings(g.gab_df if isinstance(g.gab_df, pd.DataFrame) else pd.DataFrame())
            for w in gab_w:
                w["group_key"] = g.key
                w["section"] = "갑구"
                rank_jump_warnings.append(w)

            eul_w = _find_rank_jump_warnings(g.eul_df if isinstance(g.eul_df, pd.DataFrame) else pd.DataFrame())
            for w in eul_w:
                w["group_key"] = g.key
                w["section"] = "을구"
                rank_jump_warnings.append(w)

        debug_info["rank_jump_warnings"] = rank_jump_warnings

    return groups, total_pages, debug_info


# ============================================================
# 11) Streamlit UI
# ============================================================
def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="📄", layout="wide")
    st.title(APP_TITLE)
    st.caption(f"{APP_VERSION} | Naver Table OCR(enableTableDetection) → 표제부/갑구/을구 정리")

    st.markdown(
        """
### 기능
- 네이버 **표추출 OCR(enableTableDetection)** 로 PDF 전체를 인식
- **표제부**: 표시번호/접수/소재지번/지목/면적/등기원인 및 기타사항 정리
- **갑구**: 순위번호/등기목적/접수/등기원인/권리자 및 기타사항 정리
- **을구**: 순위번호/등기목적/접수/등기원인/권리자 및 기타사항 정리
- 표제부/갑구/을구가 여러 지번으로 존재하면 **지번별로 표를 분리**해서 보여주고 엑셀 시트도 분리

> 표 추출은 네이버 콘솔 Domain에서 **‘표 추출 여부’가 ON**이어야 합니다.
"""
    )

    # 데모 접근
    with st.expander("🔐 접근(데모용)", expanded=True):
        password = st.text_input("비밀번호", type="password")
        if password != DEFAULT_PASSWORD:
            st.warning("비밀번호가 올바르지 않습니다.")
            st.stop()
        st.success("접속 완료")

    with st.sidebar:
        st.header("⚙️ API 설정")
        try:
            api_url = st.secrets["NAVER_API_URL"]
            secret_key = st.secrets["NAVER_SECRET_KEY"]
            st.success("st.secrets에서 API 정보를 불러왔습니다.")
        except Exception:
            api_url = st.text_input("NAVER_API_URL (…/general)")
            secret_key = st.text_input("NAVER_SECRET_KEY", type="password")

        st.divider()
        st.header("🧾 OCR 옵션")
        pages_per_req = st.number_input("OCR 요청당 페이지 수(<=10)", min_value=1, max_value=10, value=10, step=1)
        lang = st.selectbox("언어(lang)", options=["ko", "ja", "zh-TW", "ko,ja"], index=0)

        st.divider()
        st.header("📦 출력 옵션")
        include_raw_pyo = st.checkbox("표제부 raw 테이블 시트 포함", value=False)
        include_raw_gab = st.checkbox("갑구 raw 테이블 시트 포함", value=False)
        include_raw_eul = st.checkbox("을구 raw 테이블 시트 포함", value=False)
        merge_cells_on_raw = st.checkbox("raw 시트에서 병합셀 반영", value=True)

        st.divider()
        st.header("🧠 온톨로지(JSON-LD)")
        export_jsonld = st.checkbox("표제부+갑구+을구 JSON-LD 생성", value=False)

        st.divider()
        st.header("🧪 디버그")
        debug_extract = st.checkbox("페이지/테이블 ID 디버그 로그 표시", value=False)

    uploaded_file = st.file_uploader("📎 PDF 업로드", type=["pdf"])
    if uploaded_file is None:
        st.info("PDF를 업로드하면 시작할 수 있어요.")
        return

    file_bytes = uploaded_file.getvalue()
    file_hash = hashlib.sha256(file_bytes).hexdigest()

    if st.session_state.get("file_hash") != file_hash:
        for k in ["groups", "excel_bytes", "jsonld_obj", "jsonld_bytes", "total_pages", "debug_info"]:
            st.session_state.pop(k, None)
        st.session_state["file_hash"] = file_hash

    clicked = st.button("🚀 표제부+갑구+을구 추출 시작", disabled=not bool(api_url and secret_key))
    if clicked:
        if not api_url or not secret_key:
            st.error("API URL/SECRET을 입력하세요.")
            st.stop()

        progress = st.progress(0)
        status = st.empty()

        def progress_cb(i, total, sp, ep):
            pct = int(i / max(1, total) * 100)
            progress.progress(min(100, pct))
            if sp and ep:
                status.write(f"📄 OCR 진행 {i}/{total} (페이지 {sp}~{ep})")
            else:
                status.write("📄 마무리 중...")

        with st.spinner("OCR 및 표제부/갑구/을구 정리 중..."):
            groups, total_pages, debug_info = process_pdf(
                file_bytes,
                api_url,
                secret_key,
                pages_per_request=int(pages_per_req),
                lang=str(lang),
                progress_cb=progress_cb,
                debug=bool(debug_extract),
            )

            excel_bytes = build_registry_excel_bytes(
                groups=groups,
                include_raw_pyo=include_raw_pyo,
                include_raw_gab=include_raw_gab,
                include_raw_eul=include_raw_eul,
                merge_cells_on_raw=merge_cells_on_raw,
            )

            jsonld_obj = None
            jsonld_bytes = b""
            if export_jsonld:
                jsonld_obj = build_registry_jsonld(
                    file_name=uploaded_file.name,
                    file_hash=file_hash,
                    groups=groups,
                )
                jsonld_bytes = make_jsonld_bytes(jsonld_obj)

        st.session_state["groups"] = groups
        st.session_state["excel_bytes"] = excel_bytes
        st.session_state["jsonld_obj"] = jsonld_obj
        st.session_state["jsonld_bytes"] = jsonld_bytes
        st.session_state["total_pages"] = total_pages
        st.session_state["debug_info"] = debug_info

        progress.progress(100)
        status.write("✅ 완료")

    # 결과 표시
    groups_obj = st.session_state.get("groups")
    if groups_obj is None:
        st.info("추출을 실행하면 결과가 여기에 표시됩니다.")
        return

    groups: List[ParcelGroup] = groups_obj
    excel_bytes: bytes = st.session_state.get("excel_bytes", b"")
    jsonld_bytes: bytes = st.session_state.get("jsonld_bytes", b"")
    debug_info_obj: Dict[str, Any] = st.session_state.get("debug_info", {})

    st.divider()
    st.subheader(f"✅ 추출 결과 (지번 그룹 수: {len(groups)})")

    # 다운로드
    if excel_bytes:
        st.download_button(
            "📥 엑셀 다운로드 (지번별 시트: 표제부/갑구/을구)",
            data=excel_bytes,
            file_name=f"{uploaded_file.name}_등기정리.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel",
        )

    if export_jsonld and jsonld_bytes:
        st.download_button(
            "🧠 JSON-LD 다운로드",
            data=jsonld_bytes,
            file_name=f"{uploaded_file.name}_등기정리.jsonld",
            mime="application/ld+json",
            key="download_jsonld",
        )

    # 그룹별 표시 (탭/익스팬더)
    group_labels = [f"{g.key} (p{g.start_page}-{g.end_page})" for g in groups]

    if len(groups) <= 8:
        tabs = st.tabs(group_labels)
        for tab, g in zip(tabs, groups):
            with tab:
                _render_group(g)
    else:
        for g in groups:
            with st.expander(f"📌 {g.key} (p{g.start_page}-{g.end_page})", expanded=False):
                _render_group(g)

    if export_jsonld:
        with st.expander("🧠 JSON-LD 미리보기", expanded=False):
            st.json(st.session_state.get("jsonld_obj", {}))

    if debug_extract and isinstance(debug_info_obj, dict):
        _render_debug_info(debug_info_obj)


def _render_debug_info(debug_info: Dict[str, Any]):
    with st.expander("🧪 디버그 로그 (페이지/테이블 ID)", expanded=False):
        st.caption("보안상 API 키/원문 OCR 텍스트는 표시하지 않고, 페이지/테이블 메타 정보만 표시합니다.")

        land_pages = debug_info.get("land_registry_pages", [])
        if isinstance(land_pages, list) and len(land_pages) > 0:
            st.write(f"토지등기 포함 페이지: {land_pages}")
        else:
            st.write("토지등기 포함 페이지: 미검출")

        table_counts = debug_info.get("table_counts", {})
        if isinstance(table_counts, dict) and table_counts:
            st.write(
                "테이블 수 요약: "
                f"전체={table_counts.get('all_tables', 0)}, "
                f"토지등기 범위={table_counts.get('registry_tables', 0)}"
            )

        sec_candidates = debug_info.get("section_candidates", [])
        sec_rejected = debug_info.get("section_rejected_tables", [])
        sec_skipped = debug_info.get("section_skipped_tables", [])
        sec_ok = debug_info.get("section_accepted_tables", [])
        if any(isinstance(x, list) and x is not None for x in [sec_candidates, sec_rejected, sec_skipped, sec_ok]):
            st.write(
                "갑/을구 파이프라인 요약: "
                f"후보={len(sec_candidates) if isinstance(sec_candidates, list) else 0}, "
                f"후보탈락={len(sec_rejected) if isinstance(sec_rejected, list) else 0}, "
                f"후보후스킵={len(sec_skipped) if isinstance(sec_skipped, list) else 0}, "
                f"최종반영={len(sec_ok) if isinstance(sec_ok, list) else 0}"
            )

        page_trace = debug_info.get("land_page_trace", [])
        if isinstance(page_trace, list) and len(page_trace) > 0:
            st.markdown("#### 페이지 모드 추적")
            st.dataframe(pd.DataFrame(page_trace), use_container_width=True, hide_index=True)

        pyo_empty = debug_info.get("pyo_empty_tables", [])
        if isinstance(pyo_empty, list) and len(pyo_empty) > 0:
            st.markdown("#### 표제부 후보 중 빈 결과")
            st.dataframe(pd.DataFrame(pyo_empty), use_container_width=True, hide_index=True)

        st.markdown("#### 갑/을구 후보 탈락 테이블")
        if isinstance(sec_rejected, list) and len(sec_rejected) > 0:
            st.dataframe(pd.DataFrame(sec_rejected), use_container_width=True, hide_index=True)
        else:
            st.info("갑/을구 후보에서 탈락한 테이블이 없습니다.")

        sec_skipped = debug_info.get("section_skipped_tables", [])
        if isinstance(sec_skipped, list) and len(sec_skipped) > 0:
            st.markdown("#### 갑/을구 스킵 테이블")
            st.dataframe(pd.DataFrame(sec_skipped), use_container_width=True, hide_index=True)

        sec_ok = debug_info.get("section_accepted_tables", [])
        if isinstance(sec_ok, list) and len(sec_ok) > 0:
            st.markdown("#### 갑/을구 반영 테이블")
            st.dataframe(pd.DataFrame(sec_ok), use_container_width=True, hide_index=True)

        eul_recovery = debug_info.get("eul_gap_recovery", [])
        if isinstance(eul_recovery, list) and len(eul_recovery) > 0:
            st.markdown("#### 을구 결번 복구 로그")
            st.dataframe(pd.DataFrame(eul_recovery), use_container_width=True, hide_index=True)

        eul_fields_recovery = debug_info.get("eul_fields_recovery", [])
        if isinstance(eul_fields_recovery, list) and len(eul_fields_recovery) > 0:
            st.markdown("#### 을구 fields 라인 복구 로그")
            st.dataframe(pd.DataFrame(eul_fields_recovery), use_container_width=True, hide_index=True)

        eul_source_provenance = debug_info.get("eul_source_provenance", [])
        if isinstance(eul_source_provenance, list) and len(eul_source_provenance) > 0:
            st.markdown("#### 을구 원응답 출처 진단")
            st.dataframe(pd.DataFrame(eul_source_provenance), use_container_width=True, hide_index=True)

        jumps = debug_info.get("rank_jump_warnings", [])
        if isinstance(jumps, list) and len(jumps) > 0:
            st.markdown("#### 순위번호 점프 의심 구간")
            st.dataframe(pd.DataFrame(jumps), use_container_width=True, hide_index=True)
        else:
            st.markdown("#### 순위번호 점프 의심 구간")
            st.info("디버그 기준으로 점프 의심 구간이 발견되지 않았습니다.")


def _render_group(g: ParcelGroup):
    # 표제부
    st.markdown("#### 표제부")
    pyo_df = g.pyo_df if isinstance(g.pyo_df, pd.DataFrame) else pd.DataFrame()
    pyo_cols = ["표시번호", "접수", "소재지번", "지목", "면적", "등기원인 및 기타사항"]
    if pyo_df.empty:
        st.info("표제부 없음")
    else:
        st.dataframe(
            pyo_df[[c for c in pyo_cols if c in pyo_df.columns]],
            use_container_width=True,
            hide_index=True,
        )

    # 갑구
    st.markdown("#### 갑구")
    gab_df = g.gab_df if isinstance(g.gab_df, pd.DataFrame) else pd.DataFrame()
    gab_cols = ["순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
    if gab_df.empty:
        st.info("갑구 없음(또는 갑구 표 검출 실패)")
        st.caption(
            "팁: '갑 구' 라벨이 OCR로 안 잡혀도, 이제는 등기목적(소유권/근저당 등) 키워드로 2차 추정합니다. "
            "그래도 비면 표 자체가 tables로 검출되지 않았을 가능성이 큽니다."
        )
    else:
        st.dataframe(
            gab_df[[c for c in gab_cols if c in gab_df.columns]],
            use_container_width=True,
            hide_index=True,
        )

    # 을구
    st.markdown("#### 을구")
    eul_df = g.eul_df if isinstance(g.eul_df, pd.DataFrame) else pd.DataFrame()
    eul_cols = ["순위번호", "등기목적", "접수", "등기원인", "권리자 및 기타사항"]
    if eul_df.empty:
        st.info("을구 없음(또는 을구 표 검출 실패)")
    else:
        st.dataframe(
            eul_df[[c for c in eul_cols if c in eul_df.columns]],
            use_container_width=True,
            hide_index=True,
        )


if __name__ == "__main__":
    main()
